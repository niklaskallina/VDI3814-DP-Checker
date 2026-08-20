"""Extraktion aus Excel-Vorlagen (.xlsx/.xlsm/.xls).

Viele GA-Funktionslisten entstehen in der VDI-Excel-Vorlage. Liegt die
Originaldatei vor, ist das die genaueste Quelle ueberhaupt - jede Zelle wird
1:1 gelesen, es gibt kein Erkennungsrisiko. Auch hier ist der Anker das
Kopfzeilenpaar "Abschnitt" / "Spalte".
"""

from __future__ import annotations

import logging
from pathlib import Path

from ..models import Cell, ColumnHeader, DataPointRow, DocumentMetadata, Footnote, RowKind
from ..textutil import extract_footnote_markers, normalize, normalize_address, parse_count
from .base import ExtractionMode, RawTable, assign_sections, drop_after_footer
from .rowfilter import classify_row

log = logging.getLogger(__name__)

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xls"}

METADATA_LABELS = {
    "gewerk": "gewerk",
    "anlage": "anlage",
    "projekt": "projekt",
    "auftraggeber projekt": "auftraggeber",
    "auftraggeber": "auftraggeber",
    "planersteller": "planersteller",
    "planstand": "planstand",
    "blatt nr": "blatt_nr",
    "informationsschwerpunkt": "informationsschwerpunkt",
    "datenkommunikationsprotokoll": "protokoll",
    "datenkommunikationsprotokoll ae asp": "protokoll",
    "ausgabedatum": "datum",
    "datum": "datum",
}


def is_excel(path: str | Path) -> bool:
    return Path(path).suffix.lower() in EXCEL_SUFFIXES


def _cell_text(value) -> str:
    """Zellwert als Text; ganzzahlige Floats ohne '.0' (Excel liefert 1.0 statt 1)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_grid(path: str | Path) -> list[list[list[str]]]:
    """Liest alle Blaetter als Textraster [blatt][zeile][spalte].

    Verbundene Zellen werden auf alle beteiligten Zellen ausgerollt, damit
    Gruppenueberschriften ueber jeder zugehoerigen Spalte stehen.
    """
    path = Path(path)
    if path.suffix.lower() == ".xls":
        import xlrd  # nur fuer das alte BIFF-Format noetig

        # formatting_info=True liefert bei .xls zusaetzlich die verbundenen
        # Zellen - noetig, damit Gruppenueberschriften ueber jeder Spalte stehen.
        try:
            book = xlrd.open_workbook(path, formatting_info=True)
        except NotImplementedError:      # pragma: no cover
            book = xlrd.open_workbook(path)
        sheets = []
        for sheet in book.sheets():
            grid = [
                [_cell_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]
            for r0, r1, c0, c1 in getattr(sheet, "merged_cells", []):
                if r0 >= len(grid) or c0 >= len(grid[r0]):
                    continue
                value = grid[r0][c0]
                if not value:
                    continue
                for r in range(r0, min(r1, len(grid))):
                    for c in range(c0, min(c1, len(grid[r]))):
                        if not grid[r][c]:
                            grid[r][c] = value
            sheets.append(grid)
        return sheets

    import openpyxl

    book = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets = []
    for sheet in book.worksheets:
        grid = [
            [_cell_text(sheet.cell(r, c).value) for c in range(1, sheet.max_column + 1)]
            for r in range(1, sheet.max_row + 1)
        ]
        for merged in sheet.merged_cells.ranges:
            top, left = merged.min_row - 1, merged.min_col - 1
            if top >= len(grid) or left >= len(grid[top]):
                continue
            value = grid[top][left]
            if not value:
                continue
            for r in range(merged.min_row - 1, min(merged.max_row, len(grid))):
                for c in range(merged.min_col - 1, min(merged.max_col, len(grid[r]))):
                    if not grid[r][c]:
                        grid[r][c] = value
        sheets.append(grid)
    return sheets


def _find_label(grid: list[list[str]], *labels: str) -> tuple[int, int] | None:
    wanted = {normalize(label) for label in labels}
    for r, row in enumerate(grid):
        for c, value in enumerate(row):
            if normalize(value) in wanted:
                return r, c
    return None


def _label_row_above(grid: list[list[str]], section_row: int, first_col: int, last_col: int) -> int | None:
    """Sucht oberhalb der Abschnitt-Zeile die Zeile mit den Spaltenbezeichnungen.

    Unterscheidungsmerkmal gegenueber den Gruppenzeilen: Spaltenbezeichnungen
    sind je Spalte verschieden, Gruppenueberschriften sind verbunden und
    wiederholen sich nach dem Ausrollen ueber viele Spalten. Gezaehlt wird
    deshalb die Anzahl UNTERSCHIEDLICHER Texte.
    """
    best, best_count = None, 0
    for r in range(max(0, section_row - 8), section_row):
        values = {
            grid[r][c] for c in range(first_col, min(last_col + 1, len(grid[r])))
            if len(grid[r][c]) > 3
        }
        if len(values) >= best_count:      # bei Gleichstand die untere Zeile
            best, best_count = r, len(values)
    return best if best_count >= 3 else None


def _ffill(grid: list[list[str]], row_index: int, start: int, end: int) -> list[str]:
    """Zieht Gruppenueberschriften nach rechts durch.

    In der alten .xls-Vorlage sind Gruppenzeilen nicht verbunden, der Text
    laeuft optisch nur ueber die Nachbarzellen. Fuer die Zuordnung wird der
    zuletzt gesetzte Wert bis zur naechsten Ueberschrift fortgeschrieben.
    """
    row = list(grid[row_index]) if 0 <= row_index < len(grid) else []
    while len(row) <= end:
        row.append("")
    current = ""
    for c in range(start, end + 1):
        if row[c]:
            current = row[c]
        else:
            row[c] = current
    return row


def _value_at(grid: list[list[str]], r: int, c: int) -> str:
    if 0 <= r < len(grid) and 0 <= c < len(grid[r]):
        return grid[r][c]
    return ""


def _is_label_row(row: list[str]) -> bool:
    """Beschriftungszeile im Fussbereich (z. B. 'Datum | Inhalt | Name | Index').

    Dort stehen die Werte UNTER der Beschriftung, sonst rechts daneben.
    """
    hits = sum(1 for value in row if normalize(value.rstrip(":")) in METADATA_LABELS)
    return hits >= 2


def _extract_metadata(grid: list[list[str]]) -> DocumentMetadata:
    metadata = DocumentMetadata()
    for r, row in enumerate(grid):
        for c, value in enumerate(row):
            key = normalize(value.rstrip(":"))
            field_name = METADATA_LABELS.get(key)
            if not field_name or getattr(metadata, field_name):
                continue
            candidates = [_value_at(grid, r, c + 1), _value_at(grid, r, c + 2)]
            if _is_label_row(row):
                candidates.insert(0, _value_at(grid, r + 1, c))
            for candidate in candidates:
                if not candidate or candidate.endswith(":"):
                    continue          # das ist die naechste Beschriftung, kein Wert
                if normalize(candidate.rstrip(":")) in METADATA_LABELS:
                    continue
                setattr(metadata, field_name, candidate)
                break
    return metadata


def _extract_footnotes(grid: list[list[str]], header_row: int) -> list[Footnote]:
    """Fussnoten stehen als '1)' + Text in den Zeilen oberhalb des Tabellenkopfs.

    Die Fussnoten sind in mehreren Spaltenbloecken nebeneinander gesetzt.
    Fortsetzungszeilen werden deshalb ueber ihre SPALTE der richtigen Fussnote
    zugeordnet - sonst wandert der Text der rechten Spalte an die linke Fussnote.
    """
    footnotes: dict[str, Footnote] = {}
    anchors: list[tuple[int, int, str]] = []      # (Zeile, Textspalte, Marker)

    for r in range(0, max(0, header_row)):
        for c, value in enumerate(grid[r]):
            text = value.strip()
            if not (0 < len(text) <= 3 and text.endswith(")") and text[:-1].isdigit()):
                continue
            marker = text[:-1]
            body_col = c + 1
            body = " ".join(
                part for part in (_value_at(grid, r, c + 1), _value_at(grid, r, c + 2)) if part
            ).strip()
            footnotes.setdefault(marker, Footnote(marker=marker, text=""))
            if body:
                footnotes[marker].text = (footnotes[marker].text + " " + body).strip()
            anchors.append((r, body_col, marker))

    for r in range(0, max(0, header_row)):
        for c, value in enumerate(grid[r]):
            text = value.strip()
            if len(text) <= 10 or (len(text) <= 3 and text.endswith(")")):
                continue
            if any(row == r and abs(col - c) <= 2 for row, col, _ in anchors):
                continue                       # das ist die Ursprungszeile selbst
            candidates = [
                (r - row, marker) for row, col, marker in anchors
                if 0 < r - row <= 5 and abs(col - c) <= 2
            ]
            if not candidates:
                continue
            marker = min(candidates)[1]
            existing = footnotes[marker]
            if text not in existing.text:
                existing.text = (existing.text + " " + text).strip()

    return [footnotes[k] for k in sorted(footnotes, key=lambda m: int(m))]


def _extract_sheet(grid: list[list[str]], sheet_index: int) -> RawTable | None:
    section_pos = _find_label(grid, "Abschnitt")
    column_pos = _find_label(grid, "Spalte")
    if section_pos is None or column_pos is None:
        return None
    section_row, label_col = section_pos
    column_row, _ = column_pos

    number_row = grid[column_row]
    column_indices = [
        c for c, value in enumerate(number_row)
        if c > label_col and value.strip().isdigit()
    ]
    if len(column_indices) < 4:
        return None

    sections = [(c, normalize_address(grid[section_row][c]))
                for c in range(label_col + 1, len(grid[section_row]))
                if normalize_address(_value_at(grid, section_row, c))]

    numbers = [number_row[c] for c in column_indices]
    section_ids = assign_sections(numbers, [s for _, s in sections])
    if not section_ids:
        # Rueckfallebene: in Excel steht die Abschnittsnummer links im Block
        # (nicht zentriert) -> letzter Abschnitt links von der Spalte gilt.
        section_ids = []
        for c in column_indices:
            applicable = [s for pos, s in sections if pos <= c]
            section_ids.append(applicable[-1] if applicable else "")

    first_col, last_col = column_indices[0], column_indices[-1]
    label_row = _label_row_above(grid, section_row, first_col, last_col)
    # Gruppen-/Untergruppenzeile stehen unmittelbar ueber der Bezeichnungszeile.
    anchor = label_row if label_row is not None else section_row
    group_rows = [r for r in (anchor - 2, anchor - 1) if r >= 0]

    filled_groups = [_ffill(grid, r, first_col, last_col) for r in group_rows]

    columns: list[ColumnHeader] = []
    for index, c in enumerate(column_indices):
        label = _value_at(grid, label_row, c) if label_row is not None else ""
        group, subgroup = "", ""
        if filled_groups:
            texts = [row[c] if c < len(row) else "" for row in filled_groups]
            texts = [t for t in texts if t and t != label]
            if texts:
                group = texts[0]
                subgroup = texts[1] if len(texts) > 1 else ""
        columns.append(
            ColumnHeader(
                index=index,
                group=group,
                subgroup=subgroup,
                label=label.strip(),
                section=section_ids[index],
                column_no=number_row[c].strip(),
                footnote_markers=extract_footnote_markers(label),
                x_center=float(c),
            )
        )

    column_of_index = {c: index for index, c in enumerate(column_indices)}
    note_cols = {c for c in column_indices
                 if normalize_address(section_ids[column_of_index[c]]) in {"4", "9"}}

    rows: list[DataPointRow] = []
    for r in range(column_row + 1, len(grid)):
        row_values = grid[r]
        left_texts = [row_values[c] for c in range(0, first_col) if c < len(row_values) and row_values[c]]
        cells: list[Cell] = []
        for c in column_indices:
            raw = _value_at(grid, r, c)
            if not raw:
                continue
            count = parse_count(raw)
            if c in note_cols:
                count = None
            cells.append(Cell(column_index=column_of_index[c], raw_value=raw,
                              count=count, note="" if count is not None else raw))
        if not left_texts and not cells:
            continue

        row_no = ""
        texts = list(left_texts)
        if texts and texts[0].isdigit():
            row_no = texts.pop(0)
        qualifier = _value_at(grid, r, label_col)
        if qualifier in texts:
            texts.remove(qualifier)
        remark = " ".join(_value_at(grid, r, c) for c in sorted(note_cols)).strip()
        klartext = " ".join(texts).strip()

        has_values = any(cell.count is not None for cell in cells)
        # Nur die Beschriftung bewerten - die Zeilennummer allein macht
        # aus einer leeren Zeile keinen Datenpunkt.
        kind, reason = classify_row(klartext, has_values, remark)
        if kind is RowKind.FUSSBEREICH:
            # Unterhalb der Tabelle beginnt der Fussbereich - ab hier nichts mehr
            break
        rows.append(
            DataPointRow(
                row_no=row_no,
                klartext=klartext,
                qualifier=qualifier,
                remark=remark,
                cells=cells,
                page_index=sheet_index,
                kind=kind,
                exclusion_reason=reason,
                confidence=1.0,
            )
        )

    rows = drop_after_footer(rows)
    texts = [" ".join(value for value in row if value) for row in grid[:max(section_row + 2, 12)]]
    return RawTable(
        page_index=sheet_index,
        mode=ExtractionMode.EXCEL,
        columns=columns,
        rows=rows,
        footnotes=_extract_footnotes(grid, header_row=min(group_rows) if group_rows else section_row),
        metadata=_extract_metadata(grid),
        texts=texts + [" ".join(c.label for c in columns)],
    )


def extract_excel(path: str | Path) -> list[RawTable]:
    """Liefert je Arbeitsblatt eine RawTable (Blaetter ohne Raster entfallen)."""
    tables: list[RawTable] = []
    for index, grid in enumerate(read_grid(path)):
        try:
            table = _extract_sheet(grid, index)
        except Exception as exc:  # pragma: no cover - defensiv
            log.warning("Blatt %d in %s nicht lesbar: %s", index, path, exc)
            continue
        if table is not None:
            tables.append(table)
    return tables
