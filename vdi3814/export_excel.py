"""Excel-Export (.xlsx) mit lebenden Formeln.

Sichtbare Arbeitsblaetter - das ist der Arbeitsweg von oben nach unten:
    Übersicht       - Summen je Funktionsspalte ueber alle importierten Listen
    Schwerpunkte    - Datenpunkte und Funktionen je ASP/ISP ueber alle Listen
    Mengen je Schwerpunkt - dieselbe Sicht im Aufbau der VDI-Funktionsliste
    GA-Funktionsliste - alle Datenpunkte im Layout der Papierliste
    Kostenschätzung - Einheitspreis je Spalte, Menge per Formel aus "Übersicht",
                      Kosten = Menge * Einheitspreis, Gesamtsumme per SUM()
    Prüfung         - Abgleich mit der Zeile "Summe Funktionen" und alle bewusst
                      nicht gezaehlten Zeilen mit Begruendung

Ausgeblendete Nebenblaetter (siehe NEBENBLAETTER) - sie werden gebraucht, aber
nicht taeglich angesehen. In Excel per Rechtsklick auf einen Blattreiter ->
"Einblenden" jederzeit sichtbar zu machen:
    Spaltensummen   - flache Liste aller Funktionsspalten zum Sortieren/Filtern
    Rohdaten        - jede Einzelzelle mit Quelldatei-, Seiten- und Zeilenbezug
    Dokumente       - importierte Dateien inkl. uebersprungener Seiten/Hinweise
    Fußnoten        - Zaehlregeln der Kopfzeile je Spalte
    Projekte        - Kreuztabelle Projekt/Anlage x Spaltengruppe
    Info            - Erstellungsdatum und Eckwerte des Exports

Jedes Blatt rechnet mit Formeln statt mit Kopien: was sich aendern kann, ist
verknuepft, und jede Tabelle traegt unten eine Summenzeile. Die einzige Quelle
sind die Zeilen der "GA-Funktionsliste"; darauf verweisen "Übersicht" und
"Mengen je Schwerpunkt" per SUMIFS/COUNTIFS, darauf wiederum "Schwerpunkte",
"Spaltensummen", "Kostenschätzung", "Dokumente" und "Info". Wer in der
Funktionsliste Zeilen loescht oder ergaenzt, sieht die Aenderung ueberall
sofort. Verknuepft wird ueber ausgeblendete Schluesselspalten, weil
Dateinamen nicht eindeutig sein muessen.

Nur die Einheitspreise sind bewusst schlichte Zahlen - sie werden eingetippt.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .aggregate import SCHLUESSEL_SPALTE
from .config import SETTINGS
from .schwerpunkt import OHNE_ZUORDNUNG

# Blaetter, die zwar gebraucht werden, aber nicht auf den ersten Blick. Sie
# bleiben in der Mappe - Formeln verweisen darauf - und werden nur
# ausgeblendet, damit die Reiterleiste uebersichtlich bleibt.
NEBENBLAETTER = ("Spaltensummen", "Rohdaten", "Dokumente", "Fußnoten",
                 "Projekte", "Info")

# Titel der ausgeblendeten Schluesselspalten.
SP_SCHLUESSEL_SPALTE = "Schwerpunkt-Schlüssel"

HEADER_FILL = PatternFill("solid", fgColor="DDE5F0")
GROUP_FILL = PatternFill("solid", fgColor="C9D6E8")
THIN = Side(style="thin", color="9BA7B4")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)

# Mengen werden als Formel geschrieben. Damit die Liste trotzdem so licht
# aussieht wie die Papiervorlage, bleiben Nullen unsichtbar.
MENGE_FORMAT = "0.###;-0.###;;@"


@dataclass(frozen=True)
class _ListenBezug:
    """Wo die Datenzeilen einer "GA-Funktionsliste" im Arbeitsblatt stehen.

    Damit kann die "Übersicht" auf das Blatt verweisen, obwohl sie vorher
    geschrieben wird - die Geometrie des Layoutblatts steht vorab fest.
    """

    blatt: str
    erste_zeile: int
    letzte_zeile: int
    erste_spalte: int
    schluessel_spalte: int
    # Zweiter Schluessel: Liste + Schwerpunkt. Darueber zieht das Blatt
    # "Mengen je Schwerpunkt" genau die Zeilen eines ASP/ISP zusammen.
    sp_schluessel_spalte: int = 0

    def _bereich(self, spalte: int, fest: bool = False) -> str:
        buchstabe = get_column_letter(spalte)
        dollar = "$" if fest else ""
        return (f"'{self.blatt}'!{dollar}{buchstabe}${self.erste_zeile}:"
                f"{dollar}{buchstabe}${self.letzte_zeile}")

    def menge_formel(self, spalten_index: int, kriterium: str,
                     schluessel_spalte: int = 0) -> str:
        """Summe einer Funktionsspalte ueber alle Zeilen einer Liste."""
        spalte = schluessel_spalte or self.schluessel_spalte
        return (f"=SUMIFS({self._bereich(self.erste_spalte + spalten_index)},"
                f"{self._bereich(spalte, fest=True)},{kriterium})")

    def datenpunkt_formel(self, kriterium: str, schluessel_spalte: int = 0) -> str:
        """Anzahl der Zeilen, die zu einer Liste gehoeren."""
        spalte = schluessel_spalte or self.schluessel_spalte
        return f"=COUNTIFS({self._bereich(spalte, fest=True)},{kriterium})"


@dataclass(frozen=True)
class _UebersichtBezug:
    """Wohin die "Übersicht" verweist - fuer alle Blaetter, die daran haengen."""

    # Spaltenschluessel -> Zelle der Summenzeile ("'Übersicht'!J12")
    spalten: dict[str, str]
    # Listenschluessel ("D7") -> (Zelle Datenpunkte, Zelle Funktionen gesamt)
    dokumente: dict[str, tuple[str, str]]
    gesamt_datenpunkte: str
    gesamt_funktionen: str


@dataclass(frozen=True)
class _SchwerpunktBezug:
    """Wo die Zeilen eines Blatts "Mengen je Schwerpunkt" stehen."""

    blatt: str
    erste_zeile: int
    letzte_zeile: int
    schluessel_spalte: int          # ausgeblendet: "Schwerpunkt|Bezeichnung"
    datenpunkt_spalte: int
    gesamt_spalte: int
    schluessel: tuple[str, ...]     # welche Schwerpunkte stehen auf dem Blatt

    def _bereich(self, spalte: int) -> str:
        buchstabe = get_column_letter(spalte)
        return (f"'{self.blatt}'!${buchstabe}${self.erste_zeile}:"
                f"${buchstabe}${self.letzte_zeile}")

    def teilsumme(self, spalte: int, kriterium: str) -> str:
        return f"SUMIFS({self._bereich(spalte)},{self._bereich(self.schluessel_spalte)},{kriterium})"


def _write_frame(sheet: Worksheet, frame: pd.DataFrame, freeze: str = "A2") -> None:
    sheet.append([str(column) for column in frame.columns])
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for record in frame.itertuples(index=False):
        sheet.append([_clean(value) for value in record])
    sheet.freeze_panes = freeze
    _autosize(sheet)


def _clean(value):
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value)
    if pd.isna(value):
        return ""
    return value


def _autosize(sheet: Worksheet, maximum: int = 55) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = max((len(str(cell.value)) for cell in column_cells[:200] if cell.value is not None),
                    default=10)
        sheet.column_dimensions[letter].width = min(max(10, width + 2), maximum)


def _sheet_spaltensummen(workbook: Workbook, summary: pd.DataFrame,
                         bezuege: dict[str, str] | None = None) -> int:
    """Flache Liste aller Funktionsspalten mit ihrer Gesamtmenge.

    Ergaenzt die Uebersicht im VDI-Aufbau: hier laesst sich nach Gruppe oder
    Menge sortieren und filtern. Die Menge verweist - wo moeglich - auf die
    Summenzeile der Uebersicht, damit auch dieses Blatt mitzieht, wenn in der
    "GA-Funktionsliste" Zeilen geloescht werden.
    """
    sheet = workbook.create_sheet("Spaltensummen")
    frame = summary.rename(columns={
        "spalte_adresse": "Abschnitt.Spalte",
        "spalte_key": "Spaltenschluessel",
        "gruppe": "Gruppe",
        "untergruppe": "Untergruppe",
        "spalte": "Funktionsspalte",
        "menge": "Menge",
        "dokumente": "Dateien",
        "datenpunkte": "Datenpunkte",
        "profil": "Profil",
        "fussnoten": "Zaehlregeln (Fussnoten)",
    })
    _write_frame(sheet, frame)

    last_row = sheet.max_row
    menge_col = list(frame.columns).index("Menge") + 1
    letter = get_column_letter(menge_col)
    for offset, schluessel in enumerate(summary["spalte_key"], start=2):
        bezug = (bezuege or {}).get(schluessel)
        if bezug:
            sheet.cell(row=offset, column=menge_col, value=f"={bezug}")
    total_row = last_row + 1
    sheet.cell(row=total_row, column=1, value="Summe Funktionen").font = HEADER_FONT
    total_cell = sheet.cell(row=total_row, column=menge_col,
                            value=f"=SUM({letter}2:{letter}{last_row})")
    total_cell.font = HEADER_FONT
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=total_row, column=column).fill = TOTAL_FILL
    return menge_col


def _sheet_costs(workbook: Workbook, summary: pd.DataFrame, prices: dict[str, float],
                 menge_col: int, currency: str,
                 bezuege: dict[str, str] | None = None) -> None:
    sheet = workbook.create_sheet("Kostenschätzung")
    headers = ["Abschnitt.Spalte", "Gruppe", "Untergruppe", "Funktionsspalte",
               "Spaltenschluessel", "Menge", f"Einheitspreis [{currency}]", f"Kosten [{currency}]"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    menge_letter = get_column_letter(menge_col)
    for offset, record in enumerate(summary.itertuples(index=False), start=2):
        # Menge als Formel auf das Uebersichtsblatt - aendert sich dort etwas,
        # rechnet die Kostenschaetzung automatisch mit.
        sheet.cell(row=offset, column=1, value=record.spalte_adresse)
        sheet.cell(row=offset, column=2, value=record.gruppe)
        sheet.cell(row=offset, column=3, value=record.untergruppe)
        sheet.cell(row=offset, column=4, value=record.spalte)
        sheet.cell(row=offset, column=5, value=record.spalte_key)
        # Menge als Formel: bevorzugt auf die Summenzeile der VDI-Uebersicht,
        # sonst auf die flache Spaltenliste. So gibt es nur eine Quelle.
        bezug = (bezuege or {}).get(record.spalte_key)
        sheet.cell(row=offset, column=6,
                   value=f"={bezug}" if bezug else f"='Spaltensummen'!{menge_letter}{offset}")
        price_cell = sheet.cell(row=offset, column=7, value=float(prices.get(record.spalte_key, 0.0)))
        price_cell.number_format = "#,##0.00"
        cost_cell = sheet.cell(row=offset, column=8, value=f"=F{offset}*G{offset}")
        cost_cell.number_format = "#,##0.00"

    last_row = sheet.max_row
    total_row = last_row + 1
    sheet.cell(row=total_row, column=1, value="Gesamtkosten").font = HEADER_FONT
    sheet.cell(row=total_row, column=6, value=f"=SUM(F2:F{last_row})").font = HEADER_FONT
    total = sheet.cell(row=total_row, column=8, value=f"=SUM(H2:H{last_row})")
    total.font = HEADER_FONT
    total.number_format = "#,##0.00"
    for column in range(1, 9):
        sheet.cell(row=total_row, column=column).fill = TOTAL_FILL
    sheet.freeze_panes = "A2"
    _autosize(sheet)


def export_workbook(path: str | Path,
                    summary: pd.DataFrame,
                    raw: pd.DataFrame,
                    documents: pd.DataFrame,
                    footnotes: pd.DataFrame,
                    prices: dict[str, float] | None = None,
                    projects: pd.DataFrame | None = None,
                    currency: str | None = None,
                    layouts: dict[str, pd.DataFrame] | None = None,
                    pruefung: pd.DataFrame | None = None,
                    matrizen: dict[str, pd.DataFrame] | None = None,
                    schwerpunkte: pd.DataFrame | None = None,
                    schwerpunkt_matrizen: dict[str, pd.DataFrame] | None = None,
                    nebenblaetter_ausblenden: bool = True) -> Path:
    """Schreibt die vollstaendige Auswertung als .xlsx.

    ``nebenblaetter_ausblenden`` blendet die Blaetter aus NEBENBLAETTER aus -
    sie bleiben in der Mappe und werden weiter berechnet, stehen aber nicht
    mehr in der Reiterleiste im Weg.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    prices = prices or {}
    currency = currency or SETTINGS.currency

    workbook = Workbook()
    workbook.remove(workbook.active)

    from .profiles_loader import get_profile

    # Die Funktionslisten werden erst spaeter geschrieben, ihre Geometrie steht
    # aber schon fest. So kann die Uebersicht bereits auf sie verweisen und
    # trotzdem als erstes Blatt stehen.
    listen_titel: dict[str, str] = {}
    listen_bezug: dict[str, _ListenBezug] = {}
    for profile_id, frame in (layouts or {}).items():
        profile = get_profile(profile_id)
        if profile is None or frame is None or frame.empty:
            continue
        listen_titel[profile_id] = f"GA-Funktionsliste {profile.fassung}"[:31]
        listen_bezug[profile_id] = _bezug_funktionsliste(
            listen_titel[profile_id], len(frame), len(profile.columns))

    # Zuerst die Uebersicht im Aufbau der VDI-Funktionsliste: eine Zeile je
    # importierter Datei, darunter Summe, Einheitspreis und Kosten.
    bezuege: dict[str, str] = {}
    dokument_bezuege: dict[str, tuple[str, str]] = {}
    gesamt_datenpunkte: list[str] = []
    gesamt_funktionen: list[str] = []
    mehrere = len(matrizen or {}) > 1
    for profile_id, matrix in (matrizen or {}).items():
        profile = get_profile(profile_id)
        if profile is None or matrix is None or matrix.empty:
            continue
        titel = (f"Übersicht {profile.fassung}" if mehrere else "Übersicht")[:31]
        bezug = _sheet_uebersicht_vdi(workbook, titel, profile, matrix, prices,
                                      currency, listen_bezug.get(profile_id))
        bezuege.update(bezug.spalten)
        dokument_bezuege.update(bezug.dokumente)
        gesamt_datenpunkte.append(bezug.gesamt_datenpunkte)
        gesamt_funktionen.append(bezug.gesamt_funktionen)

    # Danach die Zuordnung zu den Automations-/Informationsschwerpunkten:
    # welcher ASP/ISP traegt wie viele Datenpunkte und Funktionen? Das Blatt
    # wird hier nur angelegt, damit es an dieser Stelle steht - gefuellt wird
    # es, sobald die Blaetter feststehen, auf die es verweist.
    schwerpunkt_blatt = (workbook.create_sheet("Schwerpunkte")
                         if schwerpunkte is not None and not schwerpunkte.empty else None)
    sp_bezuege: list[_SchwerpunktBezug] = []
    mehrere_sp = len(schwerpunkt_matrizen or {}) > 1
    for profile_id, matrix in (schwerpunkt_matrizen or {}).items():
        profile = get_profile(profile_id)
        if profile is None or matrix is None or matrix.empty:
            continue
        titel = (f"Mengen je Schwerpunkt {profile.fassung}" if mehrere_sp
                 else "Mengen je Schwerpunkt")[:31]
        sp_bezuege.append(_sheet_schwerpunkt_matrix(workbook, titel, profile, matrix,
                                                    listen_bezug.get(profile_id)))
    if schwerpunkt_blatt is not None:
        _sheet_schwerpunkte(schwerpunkt_blatt, schwerpunkte, sp_bezuege)

    # Danach die Funktionslisten im Original-Layout - dort stehen die einzelnen
    # Datenpunkte, auf denen die Mengen beruhen.
    for profile_id, frame in (layouts or {}).items():
        if profile_id not in listen_titel:
            continue
        _sheet_vdi_layout(workbook, listen_titel[profile_id], frame,
                          get_profile(profile_id), prices, currency)

    menge_col = _sheet_spaltensummen(workbook, summary, bezuege)
    _sheet_costs(workbook, summary, prices, menge_col, currency, bezuege)

    if pruefung is not None and not pruefung.empty:
        _sheet_pruefung(workbook, pruefung)
    _sheet_rohdaten(workbook, raw)
    _sheet_dokumente(workbook, documents, dokument_bezuege)
    _sheet_fussnoten(workbook, footnotes)
    if projects is not None and not projects.empty:
        _sheet_projekte(workbook, projects)

    _sheet_info(workbook, summary, raw, documents, schwerpunkte,
                gesamt_datenpunkte, gesamt_funktionen)

    if nebenblaetter_ausblenden:
        _blaetter_ausblenden(workbook, NEBENBLAETTER)

    workbook.save(path)
    return path


# --------------------------------------------------------------------------
# Nebenblaetter - ausgeblendet, aber mitgerechnet
# --------------------------------------------------------------------------

def _blaetter_ausblenden(workbook: Workbook, namen) -> None:
    """Blendet die Nebenblaetter aus - eine Mappe braucht ein sichtbares Blatt."""
    sichtbar = [name for name in workbook.sheetnames if name not in namen]
    if not sichtbar:
        return
    for name in namen:
        if name in workbook.sheetnames:
            workbook[name].sheet_state = "hidden"
    workbook.active = workbook.sheetnames.index(sichtbar[0])


def _summenzeile(sheet: Worksheet, frame: pd.DataFrame, spalten,
                 beschriftung: str = "Summe") -> int:
    """Haengt unter eine flache Tabelle eine Zeile mit SUM()-Formeln.

    ``spalten`` sind die Spaltentitel, die aufsummiert werden sollen; nicht
    vorhandene werden uebergangen.
    """
    titel = [str(spalte) for spalte in frame.columns]
    letzte_zeile = sheet.max_row
    summenzeile = letzte_zeile + 1
    sheet.cell(row=summenzeile, column=1, value=beschriftung).font = HEADER_FONT
    for name in spalten:
        if name not in titel:
            continue
        spalte = 1 + titel.index(name)
        buchstabe = get_column_letter(spalte)
        # Auch eine leere Tabelle bekommt die Formel - dann steht dort eine
        # gerechnete Null statt einer eingetragenen.
        zelle = sheet.cell(row=summenzeile, column=spalte,
                           value=f"=SUM({buchstabe}2:{buchstabe}{max(letzte_zeile, 2)})")
        zelle.font = HEADER_FONT
    for spalte in range(1, len(titel) + 1):
        sheet.cell(row=summenzeile, column=spalte).fill = TOTAL_FILL
    return summenzeile


def _sheet_pruefung(workbook: Workbook, pruefung: pd.DataFrame) -> None:
    """Nachweisblatt - mit gerechneter Differenz je Summenabgleich.

    Die Differenz steht bewusst als Formel da: wer eine Zahl der eigenen
    Zaehlung von Hand korrigiert, sieht sofort, ob der Abgleich mit der Zeile
    "Summe Funktionen" der Liste dann aufgeht.
    """
    sheet = workbook.create_sheet("Prüfung")
    _write_frame(sheet, pruefung)
    titel = [str(spalte) for spalte in pruefung.columns]
    if not {"wert_liste", "wert_eigene_zaehlung"} <= set(titel):
        _summenzeile(sheet, pruefung, ("wert_liste", "wert_eigene_zaehlung"))
        return
    liste = get_column_letter(1 + titel.index("wert_liste"))
    eigene = get_column_letter(1 + titel.index("wert_eigene_zaehlung"))
    differenz_spalte = len(titel) + 1
    kopf = sheet.cell(row=1, column=differenz_spalte, value="differenz")
    kopf.font = HEADER_FONT
    kopf.fill = HEADER_FILL
    for versatz, art in enumerate(pruefung.get("art", []), start=2):
        # Nur beim Summenabgleich ist eine Differenz aussagekraeftig; bei den
        # nicht gezaehlten Zeilen steht in "wert_liste" die Zeilensumme.
        if str(art) != "Summenabgleich":
            continue
        sheet.cell(row=versatz, column=differenz_spalte,
                   value=f"={eigene}{versatz}-{liste}{versatz}")
    sheet.column_dimensions[get_column_letter(differenz_spalte)].width = 12
    _summenzeile(sheet, pruefung.assign(differenz=0.0),
                 ("wert_liste", "wert_eigene_zaehlung", "differenz"))


def _sheet_rohdaten(workbook: Workbook, raw: pd.DataFrame) -> None:
    """Jede Einzelzelle mit Quellenangabe - die Grundlage aller Summen."""
    sheet = workbook.create_sheet("Rohdaten")
    _write_frame(sheet, raw)
    if not raw.empty:
        _summenzeile(sheet, raw, ("wert",), beschriftung="Summe Funktionen")


def _sheet_dokumente(workbook: Workbook, documents: pd.DataFrame,
                     bezuege: dict[str, tuple[str, str]] | None = None) -> None:
    """Importierte Dateien - Datenpunkte und Funktionen verweisen nach oben.

    Beide Zahlen zeigen auf die Zeile derselben Datei in der "Übersicht" und
    haengen damit an den Zeilen der "GA-Funktionsliste".
    """
    sheet = workbook.create_sheet("Dokumente")
    _write_frame(sheet, documents)
    if documents.empty:
        return
    titel = [str(spalte) for spalte in documents.columns]
    for name, stelle in (("datenpunkte", 0), ("summe_funktionen", 1)):
        if name not in titel or "dokument_id" not in titel:
            continue
        spalte = 1 + titel.index(name)
        for versatz, dokument_id in enumerate(documents["dokument_id"], start=2):
            bezug = (bezuege or {}).get(f"D{int(dokument_id)}")
            if bezug:
                sheet.cell(row=versatz, column=spalte, value=f"={bezug[stelle]}")
    _summenzeile(sheet, documents, ("seiten", "seiten_ausgewertet", "datenpunkte",
                                    "summe_funktionen"))


def _sheet_fussnoten(workbook: Workbook, footnotes: pd.DataFrame) -> None:
    """Zaehlregeln der Kopfzeile - unten die Anzahl der Fussnoten."""
    sheet = workbook.create_sheet("Fußnoten")
    _write_frame(sheet, footnotes)
    letzte_zeile = sheet.max_row
    zeile = letzte_zeile + 1
    sheet.cell(row=zeile, column=1, value="Anzahl Fußnoten").font = HEADER_FONT
    anzahl = sheet.cell(row=zeile, column=2,
                        value=f"=COUNTA(A2:A{max(letzte_zeile, 2)})")
    anzahl.font = HEADER_FONT
    for spalte in range(1, max(sheet.max_column, 2) + 1):
        sheet.cell(row=zeile, column=spalte).fill = TOTAL_FILL


def _sheet_projekte(workbook: Workbook, projects: pd.DataFrame) -> None:
    """Kreuztabelle Projekt/Anlage x Spaltengruppe.

    Die Randsummen der Kreuztabelle werden als Formeln nachgezogen, damit sich
    auch hier von Hand weiterrechnen laesst.
    """
    sheet = workbook.create_sheet("Projekte")
    _write_frame(sheet, projects)
    titel = [str(spalte) for spalte in projects.columns]
    # Erste Spalte mit Zahlen: links davon stehen Projekt und Anlage.
    erste_zahl = 1 + max((titel.index(name) for name in ("projekt", "anlage")
                          if name in titel), default=0) + 1
    letzte_zeile = sheet.max_row
    gesamt_zeile = (letzte_zeile if str(sheet.cell(letzte_zeile, 1).value) == "Gesamt"
                    else 0)
    letzte_datenzeile = (gesamt_zeile - 1) if gesamt_zeile else letzte_zeile

    if "Gesamt" in titel:
        spalte = 1 + titel.index("Gesamt")
        erste = get_column_letter(erste_zahl)
        vorletzte = get_column_letter(spalte - 1)
        for zeile in range(2, letzte_zeile + 1):
            if spalte - 1 >= erste_zahl:
                sheet.cell(row=zeile, column=spalte,
                           value=f"=SUM({erste}{zeile}:{vorletzte}{zeile})")
    if gesamt_zeile and letzte_datenzeile >= 2:
        for spalte in range(erste_zahl, len(titel) + 1):
            buchstabe = get_column_letter(spalte)
            zelle = sheet.cell(row=gesamt_zeile, column=spalte,
                               value=f"=SUM({buchstabe}2:{buchstabe}{letzte_datenzeile})")
            zelle.font = HEADER_FONT
        for spalte in range(1, len(titel) + 1):
            sheet.cell(row=gesamt_zeile, column=spalte).fill = TOTAL_FILL


def _anzahl_zeilen(workbook: Workbook, blatt: str) -> str:
    """COUNTA ueber die Datenzeilen eines Blatts - ohne dessen Summenzeile."""
    if blatt not in workbook.sheetnames:
        return ""
    letzte = workbook[blatt].max_row - 1          # letzte Zeile ist die Summe
    return f"=COUNTA('{blatt}'!A2:A{max(letzte, 2)})"


def _sheet_info(workbook: Workbook, summary: pd.DataFrame, raw: pd.DataFrame,
                documents: pd.DataFrame, schwerpunkte: pd.DataFrame | None,
                gesamt_datenpunkte: list[str], gesamt_funktionen: list[str]) -> None:
    """Eckwerte des Exports - die Zaehlwerte als Formel auf die "Übersicht"."""
    info = workbook.create_sheet("Info")
    info.append(["VDI 3814 DP-Checker"])
    info.append(["Erstellt am", datetime.now().replace(microsecond=0)])
    info.append(["Dateien", _anzahl_zeilen(workbook, "Dokumente")
                 or (int(documents.shape[0]) if not documents.empty else 0)])
    info.append(["Datenpunkte (Stand Export)",
                 "=" + "+".join(gesamt_datenpunkte) if gesamt_datenpunkte
                 else (int(raw["datenpunkt"].nunique()) if not raw.empty else 0)])
    info.append(["Funktionen gesamt (Stand Export)",
                 "=" + "+".join(gesamt_funktionen) if gesamt_funktionen
                 else (float(summary["menge"].sum()) if not summary.empty else 0.0)])
    info.append(["Schwerpunkte (ASP/ISP)",
                 _anzahl_zeilen(workbook, "Schwerpunkte")
                 or (int(schwerpunkte.shape[0]) if schwerpunkte is not None
                     and not schwerpunkte.empty else 0)])
    info.append([])
    info.append(["Hinweis", "Mengen und Kosten sind Formeln - Einheitspreise koennen "
                            "direkt in 'Kostenschaetzung' geaendert werden."])
    info.append(["Hinweis", "Zeilen in 'GA-Funktionsliste' duerfen geloescht werden: "
                            "'Uebersicht', 'Mengen je Schwerpunkt', 'Schwerpunkte', "
                            "'Spaltensummen', 'Kostenschaetzung' und 'Dokumente' "
                            "rechnen automatisch neu. Die ausgeblendeten Spalten "
                            "'Schluessel' und 'Schwerpunkt-Schluessel' stellen die "
                            "Verknuepfung her und sollten stehen bleiben."])
    info.append(["Hinweis", "Nebenblaetter (" + ", ".join(NEBENBLAETTER) + ") sind "
                            "ausgeblendet: Rechtsklick auf einen Blattreiter -> "
                            "'Einblenden'."])
    info["A1"].font = Font(bold=True, size=14)
    _autosize(info)


# --------------------------------------------------------------------------
# Blaetter zu den Automations-/Informationsschwerpunkten
# --------------------------------------------------------------------------

def _sheet_schwerpunkte(sheet: Worksheet, frame: pd.DataFrame,
                        bezuege: list[_SchwerpunktBezug] | None = None) -> None:
    """Flache Uebersicht: was entfaellt auf welchen ASP/ISP?

    Eine Zeile je Schwerpunkt mit Bezeichnung, Anzahl der Datenpunkte und
    Summe der Funktionen - die Sicht, mit der sich eine Angebotssumme auf die
    einzelnen Automationsschwerpunkte aufteilen laesst.

    Datenpunkte und Funktionen werden nicht kopiert, sondern aus den Blaettern
    "Mengen je Schwerpunkt" zusammengezaehlt - und haengen darueber an den
    Zeilen der "GA-Funktionsliste". Verknuepft wird ueber eine ausgeblendete
    Spalte aus Kennung und Bezeichnung.
    """
    anzeige = frame.rename(columns={
        "schwerpunkt": "Schwerpunkt",
        "bezeichnung": "Bezeichnung",
        "art": "Art",
        "datenpunkte": "Datenpunkte",
        "menge": "Funktionen",
        "dateien": "Dateien",
        "dokumente": "Anzahl Dateien",
    })
    _write_frame(sheet, anzeige)

    titel = list(anzeige.columns)
    schluessel_spalte = len(titel) + 1
    schluessel_buchstabe = get_column_letter(schluessel_spalte)
    verwendbar = _verknuepfbare_schwerpunkte(anzeige, bezuege)
    for versatz, record in enumerate(anzeige.to_dict("records"), start=2):
        kennung = str(record.get("Schwerpunkt", ""))
        eigener = _schwerpunkt_schluessel(kennung, str(record.get("Bezeichnung", "") or ""))
        sheet.cell(row=versatz, column=schluessel_spalte, value=eigener)
        if kennung not in verwendbar:
            # Die Zeilen dieser Kennung lassen sich nicht sauber zuordnen -
            # dann bleibt der gezaehlte Wert stehen, statt per Formel
            # womoeglich falsch zu verteilen.
            continue
        kriterium = f"${schluessel_buchstabe}{versatz}"
        for spalte_titel, spalte_von in (("Datenpunkte", "datenpunkt_spalte"),
                                         ("Funktionen", "gesamt_spalte")):
            if spalte_titel not in titel:
                continue
            teile = [bezug.teilsumme(getattr(bezug, spalte_von), kriterium)
                     for bezug in (bezuege or []) if eigener in bezug.schluessel]
            if teile:
                sheet.cell(row=versatz, column=1 + titel.index(spalte_titel),
                           value="=" + "+".join(teile))

    letzte_zeile = sheet.max_row
    summenzeile = letzte_zeile + 1
    sheet.cell(row=summenzeile, column=1, value="Summe").font = HEADER_FONT
    for spalte_titel in ("Datenpunkte", "Funktionen"):
        if spalte_titel not in titel:
            continue
        spalte = 1 + titel.index(spalte_titel)
        buchstabe = get_column_letter(spalte)
        zelle = sheet.cell(row=summenzeile, column=spalte,
                           value=f"=SUM({buchstabe}2:{buchstabe}{letzte_zeile})")
        zelle.font = HEADER_FONT
    for spalte in range(1, len(titel) + 1):
        sheet.cell(row=summenzeile, column=spalte).fill = TOTAL_FILL
    _schluessel_spalte(sheet, schluessel_spalte, letzte_kopfzeile=1,
                       titel=SP_SCHLUESSEL_SPALTE)


def _verknuepfbare_schwerpunkte(anzeige: pd.DataFrame,
                                bezuege: list[_SchwerpunktBezug] | None) -> set[str]:
    """Kennungen, deren Zeilen sich eindeutig auf die Mengenblaetter abbilden.

    Diese Uebersicht fasst je Kennung und Bezeichnung zusammen, die Blaetter
    "Mengen je Schwerpunkt" je Datei und Kennung. Solange es zu jeder Zeile
    hier genau einen Schluessel dort gibt, passt die Zuordnung. Benennt eine
    Liste denselben ASP in mehreren Zeilen unterschiedlich, faellt das
    auseinander - dann bleibt der gezaehlte Wert stehen.
    """
    if "Schwerpunkt" not in anzeige.columns:
        return set()
    hier: dict[str, set[str]] = {}
    for record in anzeige.to_dict("records"):
        kennung = str(record.get("Schwerpunkt", ""))
        hier.setdefault(kennung, set()).add(
            _schwerpunkt_schluessel(kennung, str(record.get("Bezeichnung", "") or "")))
    dort: dict[str, set[str]] = {}
    for bezug in bezuege or []:
        for schluessel in bezug.schluessel:
            dort.setdefault(schluessel.split("|", 1)[0], set()).add(schluessel)
    return {kennung for kennung, schluessel in hier.items()
            if schluessel == dort.get(kennung, set())}


def _sheet_schwerpunkt_matrix(workbook: Workbook, title: str, profile,
                              matrix: pd.DataFrame,
                              liste: _ListenBezug | None = None) -> _SchwerpunktBezug:
    """Mengen je Schwerpunkt im Aufbau der VDI-Funktionsliste.

    Wie das Uebersichtsblatt, nur ist eine Zeile hier kein Dokument, sondern
    ein Automations-/Informationsschwerpunkt. Damit ist auf einen Blick zu
    sehen, welche Funktionen in welchem ASP stecken.

    Liegt die zugehoerige "GA-Funktionsliste" mit im Export (``liste``), sind
    Mengen und Datenpunkte auch hier Formeln: sie zaehlen ueber die
    ausgeblendete Spalte "Schwerpunkt-Schlüssel" genau die Zeilen jenes ASP
    zusammen. Ein Loeschen von Zeilen wirkt sich damit sofort aus.
    """
    if SCHLUESSEL_SPALTE not in matrix.columns:
        # Ohne Listenschluessel liesse sich nicht sagen, aus welcher Liste eine
        # Zeile stammt - dann bleiben die Mengen feste Werte.
        liste = None
    sheet = workbook.create_sheet(title)
    columns = list(profile.columns)
    kopf_titel = ["Schwerpunkt", "Bezeichnung", "Datei", "Projekt", "Anlage", "Datenpunkte"]
    erste_spalte, erste_zeile = _schreibe_vdi_kopf(
        sheet, columns, kopf_titel, schluss_titel="Funktionen gesamt")
    letzte_spalte = erste_spalte + len(columns) - 1
    summe_spalte = erste_spalte + len(columns)
    # Zwei ausgeblendete Schluessel: nach unten auf die Funktionsliste,
    # nach oben fuer das Blatt "Schwerpunkte".
    listen_schluessel_spalte = summe_spalte + 1
    sp_schluessel_spalte = summe_spalte + 2
    listen_buchstabe = get_column_letter(listen_schluessel_spalte)
    datenpunkt_spalte = 1 + kopf_titel.index("Datenpunkte")
    schluessel: list[str] = []

    zeile = erste_zeile
    for record in matrix.to_dict("records"):
        kriterium = f"${listen_buchstabe}{zeile}"
        for index, titel in enumerate(kopf_titel):
            cell = sheet.cell(row=zeile, column=1 + index, value=_clean(record.get(titel, "")))
            cell.border = BOX
        sheet.cell(row=zeile, column=listen_schluessel_spalte,
                   value=_sp_schluessel(str(record.get(SCHLUESSEL_SPALTE, "")),
                                        str(record.get("Schwerpunkt", ""))))
        eigener = _schwerpunkt_schluessel(str(record.get("Schwerpunkt", "")),
                                          str(record.get("Bezeichnung", "") or ""))
        sheet.cell(row=zeile, column=sp_schluessel_spalte, value=eigener)
        schluessel.append(eigener)
        if liste is not None:
            sheet.cell(row=zeile, column=datenpunkt_spalte,
                       value=liste.datenpunkt_formel(
                           kriterium, schluessel_spalte=liste.sp_schluessel_spalte))
        for index, column in enumerate(columns):
            menge = float(record.get(column.key, 0.0) or 0.0)
            cell = sheet.cell(row=zeile, column=erste_spalte + index,
                              value=menge if menge else None)
            if liste is not None:
                cell.value = liste.menge_formel(
                    index, kriterium, schluessel_spalte=liste.sp_schluessel_spalte)
                cell.number_format = MENGE_FORMAT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BOX
        gesamt = sheet.cell(
            row=zeile, column=summe_spalte,
            value=f"=SUM({get_column_letter(erste_spalte)}{zeile}:"
                  f"{get_column_letter(letzte_spalte)}{zeile})")
        gesamt.font = HEADER_FONT
        gesamt.border = BOX
        zeile += 1

    letzte_datenzeile = zeile - 1
    summenzeile = zeile
    sheet.cell(row=summenzeile, column=1, value="Summe Funktionen").font = HEADER_FONT
    for spalte in range(1, summe_spalte + 1):
        sheet.cell(row=summenzeile, column=spalte).fill = TOTAL_FILL
    for spalte in list(range(erste_spalte, summe_spalte + 1)) + [datenpunkt_spalte]:
        buchstabe = get_column_letter(spalte)
        zelle = sheet.cell(row=summenzeile, column=spalte,
                           value=(f"=SUM({buchstabe}{erste_zeile}:{buchstabe}{letzte_datenzeile})"
                                  if letzte_datenzeile >= erste_zeile else 0))
        zelle.font = HEADER_FONT
        zelle.border = BOX

    for spalte in range(1, len(kopf_titel) + 1):
        sheet.column_dimensions[get_column_letter(spalte)].width = 14 if spalte > 2 else 24
    for spalte in range(erste_spalte, summe_spalte):
        sheet.column_dimensions[get_column_letter(spalte)].width = 6
    sheet.column_dimensions[get_column_letter(summe_spalte)].width = 14
    _schluessel_spalte(sheet, listen_schluessel_spalte, letzte_kopfzeile=3)
    _schluessel_spalte(sheet, sp_schluessel_spalte, letzte_kopfzeile=3,
                       titel=SP_SCHLUESSEL_SPALTE)

    return _SchwerpunktBezug(
        blatt=title,
        erste_zeile=erste_zeile,
        letzte_zeile=max(letzte_datenzeile, erste_zeile),
        schluessel_spalte=sp_schluessel_spalte,
        datenpunkt_spalte=datenpunkt_spalte,
        gesamt_spalte=summe_spalte,
        schluessel=tuple(schluessel),
    )


# --------------------------------------------------------------------------
# Blatt im Layout der GA-Funktionsliste
# --------------------------------------------------------------------------

# Schwerpunkt, Bezeichnung, Datei, Blatt, Zeile Nr., Datenpunkt,
# Benutzeradresse, Typ
ROW_FIELD_COUNT = 8
LAYOUT_ERSTE_DATENZEILE = 6   # darueber liegen die fuenf Kopfzeilen der Vorlage


def _bezug_funktionsliste(titel: str, zeilen: int, spalten: int) -> _ListenBezug:
    """Geometrie des Layoutblatts - berechnet, bevor es geschrieben wird.

    Rechts neben den Bemerkungen liegen zwei ausgeblendete Schluesselspalten:
    erst der Schwerpunkt-Schluessel, ganz aussen der Listen-Schluessel.
    """
    erste_spalte = ROW_FIELD_COUNT + 1
    bemerkung_spalte = erste_spalte + spalten
    return _ListenBezug(
        blatt=titel,
        erste_zeile=LAYOUT_ERSTE_DATENZEILE,
        letzte_zeile=LAYOUT_ERSTE_DATENZEILE + zeilen - 1,
        erste_spalte=erste_spalte,
        schluessel_spalte=bemerkung_spalte + 2,
        sp_schluessel_spalte=bemerkung_spalte + 1,
    )


def _sp_schluessel(listen_schluessel: str, kennung: str) -> str:
    """Schluessel einer Zeile innerhalb ihrer Liste: Liste + ASP/ISP."""
    return f"{listen_schluessel}|{kennung or OHNE_ZUORDNUNG}"


def _schwerpunkt_schluessel(kennung: str, bezeichnung: str) -> str:
    """Schluessel eines Schwerpunkts ueber alle Listen hinweg."""
    return f"{kennung or OHNE_ZUORDNUNG}|{bezeichnung or ''}"


def _schluessel_spalte(sheet: Worksheet, spalte: int, letzte_kopfzeile: int,
                      titel: str = SCHLUESSEL_SPALTE) -> None:
    """Ausgeblendete Spalte mit dem Schluessel der Liste.

    Sie traegt die Verknuepfung zwischen "GA-Funktionsliste" und "Übersicht"
    und soll beim Arbeiten mit der Liste nicht stoeren.
    """
    cell = sheet.cell(row=1, column=spalte, value=titel)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if letzte_kopfzeile > 1:
        sheet.merge_cells(start_row=1, start_column=spalte,
                          end_row=letzte_kopfzeile, end_column=spalte)
    dimension = sheet.column_dimensions[get_column_letter(spalte)]
    dimension.width = 12
    dimension.hidden = True


def _schreibe_vdi_kopf(sheet: Worksheet, columns: list, kopf_titel: list[str],
                       schluss_titel: str | None = None) -> tuple[int, int]:
    """Schreibt den Tabellenkopf im Aufbau der VDI-Vorlage.

    Zeile 1/2 Gruppe und Untergruppe (zusammengefasst), Zeile 3 die senkrechte
    Spaltenbezeichnung, Zeile 4/5 die Nummernzeilen "Abschnitt" und "Spalte".

    Rueckgabe: (erste Datenspalte, erste Datenzeile).
    """
    feld_anzahl = len(kopf_titel)
    erste_spalte = feld_anzahl + 1

    for zeile, attribut in ((1, "group"), (2, "subgroup")):
        start = erste_spalte
        for index in range(len(columns) + 1):
            aktuell = getattr(columns[index], attribut) if index < len(columns) else None
            vorher = getattr(columns[index - 1], attribut) if index else None
            if index and aktuell != vorher:
                ende = erste_spalte + index - 1
                cell = sheet.cell(row=zeile, column=start, value=vorher)
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = GROUP_FILL
                if ende > start:
                    sheet.merge_cells(start_row=zeile, start_column=start,
                                      end_row=zeile, end_column=ende)
                start = erste_spalte + index

    for index, column in enumerate(columns):
        cell = sheet.cell(row=3, column=erste_spalte + index, value=column.label)
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="bottom")
        cell.font = Font(size=8)
        cell.border = BOX
    sheet.row_dimensions[3].height = 150

    for zeile, beschriftung in ((4, "Abschnitt"), (5, "Spalte")):
        cell = sheet.cell(row=zeile, column=feld_anzahl, value=beschriftung)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="right")
    for index, column in enumerate(columns):
        abschnitt, _, nummer = column.address.rpartition(".")
        for zeile, wert in ((4, abschnitt or column.address), (5, nummer)):
            cell = sheet.cell(row=zeile, column=erste_spalte + index, value=wert)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=8, bold=True)
            cell.fill = HEADER_FILL
            cell.border = BOX

    for index, titel in enumerate(kopf_titel):
        cell = sheet.cell(row=1, column=1 + index, value=titel)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.merge_cells(start_row=1, start_column=1 + index, end_row=3, end_column=1 + index)

    if schluss_titel:
        spalte = erste_spalte + len(columns)
        cell = sheet.cell(row=1, column=spalte, value=schluss_titel)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.merge_cells(start_row=1, start_column=spalte, end_row=3, end_column=spalte)

    sheet.freeze_panes = sheet.cell(row=6, column=erste_spalte)
    return erste_spalte, 6


def _sheet_uebersicht_vdi(workbook: Workbook, title: str, profile, matrix: pd.DataFrame,
                          prices: dict[str, float], currency: str,
                          liste: _ListenBezug | None = None) -> _UebersichtBezug:
    """Uebersicht im Aufbau der VDI-Funktionsliste - eine Zeile je Datei.

    Das ist die Sicht fuer die Kalkulation: welche Mengen stecken in welcher
    Liste, was ergibt das in Summe, und was kostet es bei welchem
    Einheitspreis. Alle Summen und Kosten sind Formeln.

    Liegt die zugehoerige "GA-Funktionsliste" mit im Export (``liste``), sind
    auch die Mengen je Datei Formeln: sie zaehlen ueber die ausgeblendete
    Schluesselspalte genau die Zeilen jener Liste zusammen. Werden dort Zeilen
    geloescht, rechnet die Uebersicht sofort neu. Ohne Layoutblatt bleiben die
    Mengen feste Werte - dann gibt es nichts, worauf zu verweisen waere.

    Rueckgabe: die Bezuege auf dieses Blatt - je Spaltenschluessel die
    Summenzelle (damit die Kostenschaetzung auf dieselbe Zahl verweist statt
    sie zu kopieren) und je Liste ihre Zeile (fuer das Blatt "Dokumente").
    """
    sheet = workbook.create_sheet(title)
    columns = list(profile.columns)
    kopf_titel = ["Datei", "Projekt", "Anlage", "Schwerpunkt", "Gewerk", "Blatt", "Datenpunkte"]
    erste_spalte, erste_zeile = _schreibe_vdi_kopf(
        sheet, columns, kopf_titel, schluss_titel=f"Funktionen gesamt")
    letzte_spalte = erste_spalte + len(columns) - 1
    summe_spalte = erste_spalte + len(columns)
    schluessel_spalte = summe_spalte + 1
    schluessel_buchstabe = get_column_letter(schluessel_spalte)

    datenpunkt_spalte = 1 + kopf_titel.index("Datenpunkte")
    dokument_zeilen: dict[str, int] = {}

    zeile = erste_zeile
    for record in matrix.to_dict("records"):
        kriterium = f"${schluessel_buchstabe}{zeile}"
        dokument_zeilen[str(record.get(SCHLUESSEL_SPALTE, ""))] = zeile
        for index, titel in enumerate(kopf_titel):
            wert = record.get(titel, "")
            cell = sheet.cell(row=zeile, column=1 + index, value=_clean(wert))
            cell.border = BOX
        if liste is not None:
            sheet.cell(row=zeile, column=datenpunkt_spalte,
                       value=liste.datenpunkt_formel(kriterium))
        for index, column in enumerate(columns):
            menge = float(record.get(column.key, 0.0) or 0.0)
            cell = sheet.cell(row=zeile, column=erste_spalte + index,
                              value=menge if menge else None)
            if liste is not None:
                # Statt der Kopie die Formel auf die Funktionsliste - nur so
                # wirkt sich ein Loeschen von Zeilen dort hier aus.
                cell.value = liste.menge_formel(index, kriterium)
                cell.number_format = MENGE_FORMAT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BOX
        sheet.cell(row=zeile, column=schluessel_spalte,
                   value=record.get(SCHLUESSEL_SPALTE, ""))
        gesamt = sheet.cell(
            row=zeile, column=summe_spalte,
            value=f"=SUM({get_column_letter(erste_spalte)}{zeile}:"
                  f"{get_column_letter(letzte_spalte)}{zeile})")
        gesamt.font = HEADER_FONT
        gesamt.border = BOX
        zeile += 1

    letzte_datenzeile = zeile - 1
    summenzeile = zeile
    preiszeile = zeile + 1
    kostenzeile = zeile + 2

    beschriftung = {summenzeile: "Summe Funktionen",
                    preiszeile: f"Einheitspreis [{currency}]",
                    kostenzeile: f"Kosten [{currency}]"}
    for nummer, text in beschriftung.items():
        cell = sheet.cell(row=nummer, column=1, value=text)
        cell.font = HEADER_FONT
        for spalte in range(1, summe_spalte + 1):
            sheet.cell(row=nummer, column=spalte).fill = TOTAL_FILL

    # Auch die Datenpunkte gehoeren summiert - "Info" und "Dokumente" zaehlen
    # spaeter mit dieser Zelle weiter.
    datenpunkt_buchstabe = get_column_letter(datenpunkt_spalte)
    datenpunkt_summe = sheet.cell(row=summenzeile, column=datenpunkt_spalte)
    datenpunkt_summe.value = (
        f"=SUM({datenpunkt_buchstabe}{erste_zeile}:{datenpunkt_buchstabe}{letzte_datenzeile})"
        if letzte_datenzeile >= erste_zeile else 0)
    datenpunkt_summe.font = HEADER_FONT
    datenpunkt_summe.border = BOX

    bezuege: dict[str, str] = {}
    for index, column in enumerate(columns):
        buchstabe = get_column_letter(erste_spalte + index)
        summe = sheet.cell(row=summenzeile, column=erste_spalte + index)
        if letzte_datenzeile >= erste_zeile:
            summe.value = f"=SUM({buchstabe}{erste_zeile}:{buchstabe}{letzte_datenzeile})"
        else:
            summe.value = 0
        summe.font = HEADER_FONT
        summe.border = BOX

        preis = sheet.cell(row=preiszeile, column=erste_spalte + index,
                           value=float(prices.get(column.key, 0.0)))
        preis.number_format = "#,##0.00"
        preis.border = BOX

        kosten = sheet.cell(row=kostenzeile, column=erste_spalte + index,
                            value=f"={buchstabe}{summenzeile}*{buchstabe}{preiszeile}")
        kosten.number_format = "#,##0.00"
        kosten.border = BOX

        bezuege[column.key] = f"'{title}'!{buchstabe}{summenzeile}"

    erste = get_column_letter(erste_spalte)
    letzte = get_column_letter(letzte_spalte)
    gesamtmenge = sheet.cell(row=summenzeile, column=summe_spalte,
                             value=f"=SUM({erste}{summenzeile}:{letzte}{summenzeile})")
    gesamtmenge.font = HEADER_FONT
    gesamtkosten = sheet.cell(row=kostenzeile, column=summe_spalte,
                              value=f"=SUM({erste}{kostenzeile}:{letzte}{kostenzeile})")
    gesamtkosten.font = HEADER_FONT
    gesamtkosten.number_format = "#,##0.00"

    sheet.cell(row=preiszeile, column=summe_spalte, value="je Funktion").font = Font(italic=True, size=8)
    for spalte in range(1, len(kopf_titel) + 1):
        sheet.column_dimensions[get_column_letter(spalte)].width = 22 if spalte == 1 else 16
    for spalte in range(erste_spalte, summe_spalte + 1):
        sheet.column_dimensions[get_column_letter(spalte)].width = 6
    sheet.column_dimensions[get_column_letter(summe_spalte)].width = 14
    _schluessel_spalte(sheet, schluessel_spalte, letzte_kopfzeile=3)

    gesamt_buchstabe = get_column_letter(summe_spalte)
    return _UebersichtBezug(
        spalten=bezuege,
        dokumente={
            schluessel: (f"'{title}'!{datenpunkt_buchstabe}{nummer}",
                         f"'{title}'!{gesamt_buchstabe}{nummer}")
            for schluessel, nummer in dokument_zeilen.items() if schluessel
        },
        gesamt_datenpunkte=f"'{title}'!{datenpunkt_buchstabe}{summenzeile}",
        gesamt_funktionen=f"'{title}'!{gesamt_buchstabe}{summenzeile}",
    )


def _sheet_vdi_layout(workbook: Workbook, title: str, frame: pd.DataFrame, profile,
                      prices: dict[str, float], currency: str) -> None:
    """Schreibt die Datenpunkte im Aufbau der Original-Funktionsliste.

    Kopfbereich wie in der VDI-Vorlage: Gruppe, Untergruppe, senkrechte
    Spaltenbezeichnung sowie die Nummernzeilen "Abschnitt" und "Spalte".
    Unten die Zeile "Summe Funktionen" als Formel, darunter eine Zeile fuer
    Einheitspreise und die daraus berechneten Kosten - so laesst sich direkt
    in der Liste kalkulieren.
    """
    sheet = workbook.create_sheet(title)
    columns = list(profile.columns)
    first_data_column = ROW_FIELD_COUNT + 1
    remark_column = first_data_column + len(columns)
    sp_schluessel_spalte = remark_column + 1
    schluessel_spalte = remark_column + 2

    kopf_titel = ["Schwerpunkt", "Bezeichnung", "Datei", "Blatt", "Zeile Nr.",
                  "Datenpunkt", "Benutzeradresse", "Typ"]

    # --- Zeile 1-2: Gruppe und Untergruppe, jeweils zusammengefasst ---
    for zeile, attribut in ((1, "group"), (2, "subgroup")):
        start = first_data_column
        for index in range(len(columns) + 1):
            aktuell = getattr(columns[index], attribut) if index < len(columns) else None
            vorher = getattr(columns[index - 1], attribut) if index else None
            if index and aktuell != vorher:
                ende = first_data_column + index - 1
                cell = sheet.cell(row=zeile, column=start, value=vorher)
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = GROUP_FILL
                if ende > start:
                    sheet.merge_cells(start_row=zeile, start_column=start,
                                      end_row=zeile, end_column=ende)
                start = first_data_column + index

    # --- Zeile 3: Spaltenbezeichnung, senkrecht wie im Original ---
    for index, column in enumerate(columns):
        cell = sheet.cell(row=3, column=first_data_column + index, value=column.label)
        cell.alignment = Alignment(textRotation=90, horizontal="center", vertical="bottom")
        cell.font = Font(size=8)
        cell.border = BOX
    sheet.row_dimensions[3].height = 150

    # --- Zeile 4/5: Abschnitt und Spalte ---
    for zeile, beschriftung in ((4, "Abschnitt"), (5, "Spalte")):
        cell = sheet.cell(row=zeile, column=ROW_FIELD_COUNT, value=beschriftung)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="right")
    for index, column in enumerate(columns):
        abschnitt, _, nummer = column.address.rpartition(".")
        for zeile, wert in ((4, abschnitt or column.address), (5, nummer)):
            cell = sheet.cell(row=zeile, column=first_data_column + index, value=wert)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=8, bold=True)
            cell.fill = HEADER_FILL
            cell.border = BOX

    # --- Beschriftung der Zeilenfelder: ueber die Kopfzeilen 1-3 zusammengefasst,
    #     damit die Nummernzeilen "Abschnitt"/"Spalte" frei bleiben (wie im Original) ---
    for index, titel in enumerate(kopf_titel):
        cell = sheet.cell(row=1, column=1 + index, value=titel)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.merge_cells(start_row=1, start_column=1 + index, end_row=3, end_column=1 + index)
    bemerkung = sheet.cell(row=1, column=remark_column, value="Bemerkungen")
    bemerkung.font = HEADER_FONT
    bemerkung.fill = HEADER_FILL
    bemerkung.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.merge_cells(start_row=1, start_column=remark_column, end_row=3, end_column=remark_column)

    # --- Datenzeilen ---
    erste_datenzeile = LAYOUT_ERSTE_DATENZEILE
    zeile = erste_datenzeile
    for record in frame.to_dict("records"):
        for index, titel in enumerate(kopf_titel):
            sheet.cell(row=zeile, column=1 + index, value=record.get(titel, ""))
        for index, column in enumerate(columns):
            wert = record.get(column.label)
            if wert is None or (isinstance(wert, float) and pd.isna(wert)):
                continue
            zelle = sheet.cell(row=zeile, column=first_data_column + index, value=wert)
            zelle.alignment = Alignment(horizontal="center")
        sheet.cell(row=zeile, column=remark_column, value=record.get("Bemerkungen", ""))
        # Traegt die Zeile ihrer Liste zu - darueber summiert die Uebersicht.
        listen_schluessel = record.get(SCHLUESSEL_SPALTE, "")
        sheet.cell(row=zeile, column=schluessel_spalte, value=listen_schluessel)
        # Und ihrem Schwerpunkt innerhalb dieser Liste - darueber summiert das
        # Blatt "Mengen je Schwerpunkt".
        sheet.cell(row=zeile, column=sp_schluessel_spalte,
                   value=_sp_schluessel(listen_schluessel, record.get("Schwerpunkt", "")))
        zeile += 1
    letzte_datenzeile = zeile - 1

    # --- Summenzeile als Formel ---
    summenzeile = zeile
    sheet.cell(row=summenzeile, column=1, value="Summe Funktionen").font = HEADER_FONT
    for index in range(len(columns)):
        spalte = first_data_column + index
        buchstabe = get_column_letter(spalte)
        formel = (f"=SUM({buchstabe}{erste_datenzeile}:{buchstabe}{letzte_datenzeile})"
                  if letzte_datenzeile >= erste_datenzeile else 0)
        zelle = sheet.cell(row=summenzeile, column=spalte, value=formel)
        zelle.font = HEADER_FONT
        zelle.fill = TOTAL_FILL
    gesamt_spalte = get_column_letter(remark_column)
    sheet.cell(row=summenzeile, column=remark_column,
               value=f"=SUM({get_column_letter(first_data_column)}{summenzeile}:"
                     f"{get_column_letter(remark_column - 1)}{summenzeile})").font = HEADER_FONT

    # --- Einheitspreise und Kosten ---
    preiszeile = summenzeile + 1
    kostenzeile = summenzeile + 2
    sheet.cell(row=preiszeile, column=1, value=f"Einheitspreis [{currency}]").font = HEADER_FONT
    sheet.cell(row=kostenzeile, column=1, value=f"Kosten [{currency}]").font = HEADER_FONT
    for index, column in enumerate(columns):
        spalte = first_data_column + index
        buchstabe = get_column_letter(spalte)
        preis = sheet.cell(row=preiszeile, column=spalte,
                           value=float(prices.get(column.key, 0.0)))
        preis.number_format = "#,##0.00"
        kosten = sheet.cell(row=kostenzeile, column=spalte,
                            value=f"={buchstabe}{summenzeile}*{buchstabe}{preiszeile}")
        kosten.number_format = "#,##0.00"
    gesamt = sheet.cell(
        row=kostenzeile, column=remark_column,
        value=f"=SUM({get_column_letter(first_data_column)}{kostenzeile}:"
              f"{get_column_letter(remark_column - 1)}{kostenzeile})")
    gesamt.font = HEADER_FONT
    gesamt.number_format = "#,##0.00"
    sheet.cell(row=kostenzeile - 1, column=remark_column).fill = TOTAL_FILL
    for spalte in range(1, remark_column + 1):
        sheet.cell(row=summenzeile, column=spalte).fill = TOTAL_FILL

    # --- Darstellung ---
    sheet.freeze_panes = sheet.cell(row=erste_datenzeile, column=first_data_column)
    for index in range(len(columns)):
        sheet.column_dimensions[get_column_letter(first_data_column + index)].width = 4.5
    for index, breite in enumerate((10, 24, 26, 7, 8, 34, 22, 12), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = breite
    sheet.column_dimensions[gesamt_spalte].width = 40
    _schluessel_spalte(sheet, sp_schluessel_spalte, letzte_kopfzeile=3,
                       titel=SP_SCHLUESSEL_SPALTE)
    _schluessel_spalte(sheet, schluessel_spalte, letzte_kopfzeile=3)
