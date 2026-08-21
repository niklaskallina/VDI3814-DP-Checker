"""Zuordnung der Datenpunkte zu Automations-/Informationsschwerpunkten.

Ein Plansatz enthaelt in aller Regel mehrere Schwerpunkte (ASP01, ASP02, ...).
Fuer die Kalkulation muss nachvollziehbar sein, wie viele Datenpunkte und
Funktionen auf welchen davon entfallen - genau das pruefen diese Tests.
"""

import openpyxl
import pytest
from sqlalchemy.orm import Session

from vdi3814 import aggregate, db, export_excel, schwerpunkt
from vdi3814.models import Cell, DataPointRow, RowKind
from vdi3814.pipeline import process_file


@pytest.fixture()
def session(tmp_path):
    engine = db.make_engine(tmp_path / "test.sqlite3")
    with Session(engine) as active:
        yield active


def _zeile(label, werte=None, kind=RowKind.DATEN, seite=0, bas=""):
    return DataPointRow(
        klartext=label, bas=bas, kind=kind, page_index=seite,
        cells=[Cell(column_index=i, raw_value=str(v), count=float(v))
               for i, v in (werte or {}).items()],
    )


# --------------------------------------------------------------------------
# Erkennung im Text
# --------------------------------------------------------------------------

@pytest.mark.parametrize("text, kennung, bezeichnung", [
    ("ASP01 Heizungstechnik 2.UG", "ASP01", "Heizungstechnik 2.UG"),
    ("ASP 2 Kaeltetechnik", "ASP02", "Kaeltetechnik"),
    ("asp-nr. 7 Technikzentrale", "ASP07", "Technikzentrale"),
    ("Informationsschwerpunkt: ISP04 Kaelte", "ISP04", "Kaelte"),
    ("Automationsschwerpunkt 3 Lueftung", "ASP03", "Lueftung"),
    ("Anlage: ASP03 Lueftungstechnik", "ASP03", "Lueftungstechnik"),
    ("RLT 01 ASP03", "ASP03", "RLT 01"),
])
def test_kennung_und_bezeichnung_werden_gelesen(text, kennung, bezeichnung):
    gefunden = schwerpunkt.parse(text)
    assert gefunden is not None
    assert gefunden.kennung == kennung
    assert gefunden.bezeichnung == bezeichnung


@pytest.mark.parametrize("text", ["Aspiration Luft", "Wasp", "Zulufttemperatur",
                                  "Raspberry", "", None])
def test_aehnliche_woerter_sind_kein_schwerpunkt(text):
    """Sonst wuerde jede Beschriftung mit "asp" einen Schwerpunkt erfinden."""
    assert schwerpunkt.parse(text) is None


# --------------------------------------------------------------------------
# Zuordnung der Zeilen
# --------------------------------------------------------------------------

def test_bezeichnung_aus_der_zeile_darunter():
    """Auf Uebersichtsblaettern steht die Bezeichnung unter der Kennung."""
    rows = [
        _zeile("ASP01", {0: 15, 1: 10}),
        _zeile("Heizungstechnik 2.UG"),          # zweite Zeile derselben Zelle
        _zeile("ASP02", {0: 22, 1: 17}),
        _zeile("Kaeltetechnik 2.UG"),
    ]
    schwerpunkt.zuordnen(rows)
    assert [r.schwerpunkt for r in rows] == ["ASP01", "ASP01", "ASP02", "ASP02"]
    assert rows[0].schwerpunkt_text == "Heizungstechnik 2.UG"
    assert rows[2].schwerpunkt_text == "Kaeltetechnik 2.UG"


def test_zwischenueberschrift_gilt_fuer_die_folgenden_zeilen():
    rows = [
        _zeile("ASP01 Heizung"),
        _zeile("Vorlauftemperatur", {0: 1}),
        _zeile("Ruecklauftemperatur", {0: 1}),
        _zeile("ASP02 Lueftung"),
        _zeile("Zulufttemperatur", {0: 1}),
    ]
    schwerpunkt.zuordnen(rows)
    assert [r.schwerpunkt for r in rows] == ["ASP01", "ASP01", "ASP01", "ASP02", "ASP02"]


def test_ein_einzelner_treffer_gilt_fuer_die_ganze_seite():
    """Steht der ASP erst mitten auf der Seite, gehoeren die Zeilen davor dazu."""
    rows = [
        _zeile("Vorlauftemperatur", {0: 1}),
        _zeile("Pumpe *ASP05*H01*PU", {0: 1}, bas="*ASP05*H01*PU"),
        _zeile("Ventil", {0: 1}),
    ]
    schwerpunkt.zuordnen(rows)
    assert {r.schwerpunkt for r in rows} == {"ASP05"}


def test_zweite_seite_faellt_auf_ihre_eigene_vorgabe_zurueck():
    rows = [
        _zeile("ASP01 Heizung", {0: 1}, seite=0),
        _zeile("Vorlauftemperatur", {0: 1}, seite=1),
    ]
    schwerpunkt.zuordnen(rows, vorgaben={1: schwerpunkt.Schwerpunkt("ASP09", "Kaelte")})
    assert [r.schwerpunkt for r in rows] == ["ASP01", "ASP09"]


def test_summenzeile_gehoert_nicht_zum_zuletzt_genannten_asp():
    """Sonst stuende die Seitensumme faelschlich beim letzten Schwerpunkt."""
    rows = [
        _zeile("ASP01 Heizung", {0: 5}),
        _zeile("ASP02 Lueftung", {0: 7}),
        _zeile("Summe Funktionen", {0: 12}, kind=RowKind.SUMME),
    ]
    schwerpunkt.zuordnen(rows)
    assert rows[2].schwerpunkt == ""


@pytest.mark.parametrize("kennung, bezeichnung", [
    ("ASP01", "Heizungstechnik 2.UG"),
    ("ISP04", "Kaelte"),
    ("ASP", "Heizzentrale"),
    ("ASP12", ""),
])
def test_anzeige_laesst_sich_wieder_einlesen(kennung, bezeichnung):
    """Die Korrekturspalte der Oberflaeche zeigt und liest denselben Text."""
    zeile = _zeile("egal")
    zeile.schwerpunkt, zeile.schwerpunkt_text = kennung, bezeichnung
    gefunden = schwerpunkt.parse(zeile.schwerpunkt_anzeige())
    assert (gefunden.kennung, gefunden.bezeichnung) == (kennung, bezeichnung)


# --------------------------------------------------------------------------
# Ganze Dateien
# --------------------------------------------------------------------------

def test_uebersichtsblatt_weist_jeden_asp_einzeln_aus(samples):
    """Der Fall aus den Plansaetzen: drei ASP auf einem Blatt."""
    ergebnis = process_file(samples["asp"])
    aufteilung = ergebnis.schwerpunkte()

    assert [e["schwerpunkt"] for e in aufteilung] == ["ASP01", "ASP02", "ASP03"]
    assert [e["bezeichnung"] for e in aufteilung] == [
        "Heizungstechnik 2.UG", "Kaeltetechnik 2.UG", "Lueftungstechnik 2.UG"]
    assert [e["datenpunkte"] for e in aufteilung] == [1, 1, 1]
    # Die Summen der Schwerpunkte ergeben zusammen die Gesamtmenge der Datei
    assert sum(e["funktionen"] for e in aufteilung) == ergebnis.grand_total()


def test_asp_aus_dem_fussbereich_gilt_fuer_die_ganze_liste(samples):
    """Steht der ASP nur im Fussbereich, tragen ihn trotzdem alle Zeilen."""
    ergebnis = process_file(samples["alt"])
    assert ergebnis.metadata.automationsschwerpunkt == "01"
    assert {r.schwerpunkt for r in ergebnis.counted_rows()} == {"ASP01"}
    # Ohne eigene Bezeichnung wird die Anlagenangabe des Blattes genommen
    assert ergebnis.schwerpunkte()[0]["bezeichnung"] == "RLT 01 Lueftung Buerotrakt"


def test_asp_aus_der_anlagenangabe(samples):
    ergebnis = process_file(samples["neu"])
    assert {r.schwerpunkt for r in ergebnis.counted_rows()} == {"ASP02"}


def _liste_ohne_schwerpunkt(path):
    """Kleine Excel-Funktionsliste, die nirgends einen ASP/ISP nennt."""
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet["D2"] = "1. Ein-/Ausgabefunktionen"
    sheet["D3"] = "Physikalische"
    sheet["D4"], sheet["E4"] = "Analoge Eingabe (AI)", "Binäre Eingabe (BI)"
    sheet["F4"], sheet["G4"] = "Analoge Ausgabe (AO)", "Binäre Ausgabe (BO)"
    sheet["C5"], sheet["D5"] = "Abschnitt ", "1.1."
    sheet["C6"] = "Spalte "
    sheet["D6"], sheet["E6"], sheet["F6"], sheet["G6"] = "1", "2", "3", "4"
    sheet["A7"], sheet["B7"], sheet["D7"] = "1", "Vorlauftemperatur", 1
    sheet["A8"], sheet["B8"], sheet["E8"] = "2", "Pumpe", 2
    book.save(path)
    return path


def test_fehlender_schwerpunkt_wird_gemeldet(tmp_path):
    """Ohne Angabe darf nichts geraten werden - stattdessen ein Hinweis."""
    ergebnis = process_file(_liste_ohne_schwerpunkt(tmp_path / "ohne_asp.xlsx"))

    assert [r.schwerpunkt for r in ergebnis.counted_rows()] == ["", ""]
    assert any("ASP/ISP" in hinweis for hinweis in ergebnis.warnings)
    aufteilung = ergebnis.schwerpunkte()
    assert [e["schwerpunkt"] for e in aufteilung] == [schwerpunkt.OHNE_ZUORDNUNG]
    assert aufteilung[0]["datenpunkte"] == 2


def test_mehrere_schwerpunkte_sind_kein_hinweis(samples):
    """Ein Plansatz mit ASP01-ASP03 ist der Normalfall, keine Auffaelligkeit."""
    ergebnis = process_file(samples["asp"])
    assert not any("ASP/ISP" in hinweis for hinweis in ergebnis.warnings)


def test_schwerpunkt_uebersteht_datenbank_und_auswertung(session, samples):
    db.save_document(session, process_file(samples["asp"]))
    frame = aggregate.raw_dataframe(session)

    uebersicht = aggregate.schwerpunkt_summary(frame)
    assert list(uebersicht["schwerpunkt"]) == ["ASP01", "ASP02", "ASP03"]
    assert list(uebersicht["art"]) == ["ASP", "ASP", "ASP"]
    assert list(uebersicht["bezeichnung"]) == [
        "Heizungstechnik 2.UG", "Kaeltetechnik 2.UG", "Lueftungstechnik 2.UG"]
    assert uebersicht["menge"].sum() == frame["wert"].sum()
    assert uebersicht["datenpunkte"].sum() == frame["zeile_id"].nunique()

    # Dieselben Mengen im Aufbau der VDI-Liste
    matrix = aggregate.schwerpunkt_matrix(session, "vdi3814_alt")
    assert list(matrix["Schwerpunkt"]) == ["ASP01", "ASP02", "ASP03"]
    spalten = [c for c in matrix.columns if c.startswith("A_")]
    assert matrix[spalten].to_numpy().sum() == pytest.approx(frame["wert"].sum())

    # Zurueckgeladen muss die Zuordnung unveraendert sein
    dokument = db.list_documents(session)[0]
    zurueck = db.load_document(session, dokument.id)
    assert [e["schwerpunkt"] for e in zurueck.schwerpunkte()] == ["ASP01", "ASP02", "ASP03"]


def test_dokumentuebersicht_nennt_die_schwerpunkte(session, samples):
    db.save_document(session, process_file(samples["asp"]))
    dokumente = aggregate.documents_frame(session)
    assert dokumente.loc[0, "schwerpunkte"] == "ASP01, ASP02, ASP03"


def test_export_enthaelt_die_schwerpunkte(session, samples, tmp_path):
    db.save_document(session, process_file(samples["asp"]))
    roh = aggregate.raw_dataframe(session)
    ziel = export_excel.export_workbook(
        tmp_path / "Auswertung.xlsx",
        summary=aggregate.column_summary(roh),
        raw=roh,
        documents=aggregate.documents_frame(session),
        footnotes=aggregate.footnotes_frame(session),
        matrizen={p: aggregate.datei_spalten_matrix(session, p)
                  for p in aggregate.profiles_in_use(session)},
        layouts={p: aggregate.vdi_layout_frame(session, p)
                 for p in aggregate.profiles_in_use(session)},
        schwerpunkte=aggregate.schwerpunkt_summary(roh),
        schwerpunkt_matrizen={p: aggregate.schwerpunkt_matrix(session, p)
                              for p in aggregate.profiles_in_use(session)},
    )
    buch = openpyxl.load_workbook(ziel)
    assert "Schwerpunkte" in buch.sheetnames
    blatt = buch["Schwerpunkte"]
    kopf = [zelle.value for zelle in blatt[1]]
    assert kopf[:2] == ["Schwerpunkt", "Bezeichnung"]
    kennungen = [blatt.cell(z, 1).value for z in range(2, blatt.max_row + 1)]
    assert kennungen[:3] == ["ASP01", "ASP02", "ASP03"]
    assert kennungen[-1] == "Summe"

    # Aufbau der VDI-Liste, aber eine Zeile je Schwerpunkt
    matrix = buch["Mengen je Schwerpunkt"]
    assert [matrix.cell(z, 1).value for z in range(6, 9)] == ["ASP01", "ASP02", "ASP03"]
    assert matrix.cell(9, 1).value == "Summe Funktionen"

    # Auch die Datenpunktliste im Original-Layout nennt den Schwerpunkt
    layout = [n for n in buch.sheetnames if n.startswith("GA-Funktionsliste")][0]
    assert buch[layout].cell(1, 1).value == "Schwerpunkt"
    assert buch[layout].cell(6, 1).value == "ASP01"
