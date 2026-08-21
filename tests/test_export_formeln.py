"""Abnahme des Excel-Exports.

Der Export ist das Arbeitsergebnis: dort werden Mengen abgelesen und
Einheitspreise eingetragen. Diese Tests pruefen deshalb nicht nur, dass
Formeln dastehen, sondern rechnen sie nach und vergleichen mit der Datenbank.
"""

import copy

import openpyxl
import pytest
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from vdi3814 import aggregate, db, export_excel
from vdi3814.pipeline import process_file
from vdi3814.profiles_loader import get_profile

PREISE = {"A_1_1": 45.0, "A_1_5": 80.0, "A_8_1": 12.5}


def _erste_datenspalte(blatt) -> int:
    """Erste Funktionsspalte des VDI-Kopfes.

    Links davon stehen die Zeilenfelder (Datei, Projekt, Anlage, Schwerpunkt,
    ...); ihre Anzahl darf sich aendern, ohne dass die Tests brechen.
    """
    for spalte in range(1, blatt.max_column + 1):
        if blatt.cell(4, spalte).value == "Abschnitt":
            return spalte + 1
    raise AssertionError("Kopfzeile 'Abschnitt' nicht gefunden")


@pytest.fixture(scope="module")
def export(tmp_path_factory, samples):
    """Importiert beide Beispiellisten und schreibt einen echten Export."""
    verzeichnis = tmp_path_factory.mktemp("export")
    engine = db.make_engine(verzeichnis / "test.sqlite3")
    with Session(engine) as session:
        for name in ("alt", "neu"):
            db.save_document(session, process_file(samples[name]))
        db.set_unit_prices(session, PREISE)
        roh = aggregate.raw_dataframe(session)
        ziel = export_excel.export_workbook(
            verzeichnis / "Auswertung.xlsx",
            summary=aggregate.column_summary(roh),
            raw=roh,
            documents=aggregate.documents_frame(session),
            footnotes=aggregate.footnotes_frame(session),
            prices=db.get_unit_prices(session),
            projects=aggregate.pivot_projects(roh),
            layouts={p: aggregate.vdi_layout_frame(session, p)
                     for p in aggregate.profiles_in_use(session)},
            matrizen={p: aggregate.datei_spalten_matrix(session, p)
                      for p in aggregate.profiles_in_use(session)},
            pruefung=aggregate.pruefbericht(session),
            schwerpunkte=aggregate.schwerpunkt_summary(roh),
            schwerpunkt_matrizen={p: aggregate.schwerpunkt_matrix(session, p)
                                  for p in aggregate.profiles_in_use(session)},
        )
        mengen = roh.groupby("spalte_key")["wert"].sum().to_dict()
    return ziel, mengen


def test_uebersicht_ist_ein_vdi_blatt_mit_einer_zeile_je_datei(export):
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    blaetter = [name for name in buch.sheetnames if name.startswith("Übersicht")]
    assert blaetter, "es muss eine Uebersicht im VDI-Aufbau geben"

    blatt = buch[blaetter[0]]
    # Kopfaufbau wie in der Vorlage
    assert blatt.cell(1, 1).value == "Datei"
    erste = _erste_datenspalte(blatt)
    assert blatt.cell(4, erste - 1).value == "Abschnitt"
    assert blatt.cell(5, erste - 1).value == "Spalte"
    # Nummernzeilen ergeben zusammen die Spaltenadresse
    assert f"{blatt.cell(4, erste).value}.{blatt.cell(5, erste).value}" == "1.1"

    # Je Datei genau eine Zeile, darunter Summe / Einheitspreis / Kosten
    beschriftungen = [blatt.cell(z, 1).value for z in range(6, blatt.max_row + 1)]
    assert "Summe Funktionen" in beschriftungen
    assert any(str(b).startswith("Einheitspreis") for b in beschriftungen)
    assert any(str(b).startswith("Kosten") for b in beschriftungen)
    dateien = [b for b in beschriftungen if str(b).endswith(".pdf")]
    assert len(dateien) == 1, "je Fassung eine Beispieldatei"


def test_mengen_und_kosten_sind_formeln(export):
    """Ohne Formeln kann in Excel nicht weitergerechnet werden."""
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    blatt = buch[[n for n in buch.sheetnames if n.startswith("Übersicht")][0]]
    summenzeile = blatt.max_row - 2
    kostenzeile = blatt.max_row
    erste = _erste_datenspalte(blatt)

    assert str(blatt.cell(summenzeile, erste).value).startswith("=SUM(")
    assert str(blatt.cell(kostenzeile, erste).value).startswith(f"={get_column_letter(erste)}")
    # Der Einheitspreis muss eine schlichte Zahl bleiben - er wird eingetippt
    assert isinstance(blatt.cell(summenzeile + 1, erste).value, (int, float))

    kosten = buch["Kostenschätzung"]
    assert str(kosten.cell(2, 6).value).startswith("='Übersicht"), \
        "die Menge muss auf die Uebersicht verweisen, nicht kopiert sein"
    assert kosten.cell(2, 8).value == "=F2*G2"


def test_formeln_rechnen_richtig(export):
    """Formeln werden ausgewertet und gegen die Datenbank geprueft."""
    formulas = pytest.importorskip("formulas", reason="Formelrechner nicht installiert")
    import logging
    import warnings

    warnings.filterwarnings("ignore")
    logging.disable(logging.WARNING)

    ziel, mengen = export
    modell = formulas.ExcelModel().loads(str(ziel)).finish()
    loesung = modell.calculate()
    werte = {}
    for schluessel, wert in loesung.items():
        teile = schluessel.split("!")
        if len(teile) < 2:
            continue
        blattname = teile[-2].split("]")[-1].strip("'").upper()
        try:
            werte[(blattname, teile[-1].strip("'"))] = wert.value[0, 0]
        except Exception:
            continue

    buch = openpyxl.load_workbook(ziel)
    name = [n for n in buch.sheetnames if n.startswith("Übersicht")][0]
    blatt = buch[name]
    summenzeile = blatt.max_row - 2
    kostenzeile = blatt.max_row
    profil = get_profile("vdi3814_alt" if "alt" in name or len(buch.sheetnames) else "vdi3814_alt")

    geprueft = 0
    for spalte in range(_erste_datenspalte(blatt), blatt.max_column):
        adresse = f"{blatt.cell(4, spalte).value}.{blatt.cell(5, spalte).value}"
        eintrag = profil.column_by_address(adresse)
        if eintrag is None:
            continue
        buchstabe = get_column_letter(spalte)
        summe = werte.get((name.upper(), f"{buchstabe}{summenzeile}"))
        if summe is None:
            continue
        erwartet = mengen.get(eintrag.key, 0.0)
        assert float(summe) == pytest.approx(erwartet), f"Spalte {adresse}"

        kosten = werte.get((name.upper(), f"{buchstabe}{kostenzeile}"))
        assert float(kosten) == pytest.approx(erwartet * PREISE.get(eintrag.key, 0.0)), \
            f"Kosten in Spalte {adresse}"
        geprueft += 1
    assert geprueft >= 10, "es muessen mehrere Spalten geprueft worden sein"


def test_zeilensumme_je_datei_stimmt(export):
    """Die Zeilensumme einer Datei muss deren Gesamtmenge sein."""
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    name = [n for n in buch.sheetnames if n.startswith("Übersicht")][0]
    blatt = buch[name]
    # Ganz rechts liegt die ausgeblendete Schluesselspalte, davor "Funktionen gesamt".
    gesamt_spalte = blatt.max_column - 1
    letzte = get_column_letter(gesamt_spalte)
    erste = _erste_datenspalte(blatt)

    werte = _rechne(ziel)

    zeile = 6
    while str(blatt.cell(zeile, 1).value or "").endswith(".pdf"):
        einzeln = sum(_zahl(werte, name, spalte, zeile)
                      for spalte in range(erste, gesamt_spalte))
        gesamt = werte.get((name.upper(), f"{letzte}{zeile}"))
        assert float(gesamt) == pytest.approx(einzeln)
        zeile += 1
    assert zeile > 6, "es muss mindestens eine Dateizeile geben"


def test_uebersicht_haengt_an_der_ga_funktionsliste(export):
    """Ohne Verknuepfung wuerde ein Loeschen von Zeilen unbemerkt bleiben."""
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    name = [n for n in buch.sheetnames if n.startswith("Übersicht")][0]
    blatt = buch[name]
    liste = [n for n in buch.sheetnames if n.startswith("GA-Funktionsliste")][0]
    erste = _erste_datenspalte(blatt)

    # Verknuepft wird ueber einen Schluessel je Liste - Dateinamen sind nicht
    # eindeutig. Die Spalte steht ganz rechts und ist ausgeblendet.
    schluessel = blatt.max_column
    assert blatt.cell(1, schluessel).value == "Schlüssel"
    assert blatt.column_dimensions[get_column_letter(schluessel)].hidden
    assert blatt.cell(6, schluessel).value == buch[liste].cell(6, buch[liste].max_column).value

    # Mengen und Datenpunkte werden gezaehlt, nicht kopiert.
    menge = str(blatt.cell(6, erste).value)
    assert menge.startswith("=SUMIFS("), menge
    assert f"'{liste}'!" in menge
    assert str(blatt.cell(6, erste - 1).value).startswith("=COUNTIFS(")

    # Auch die flache Spaltenliste zieht die Menge aus der Uebersicht.
    assert str(buch["Spaltensummen"].cell(2, 7).value).startswith("='Übersicht")


def test_geloeschte_zeile_wirkt_bis_in_die_kosten(export, tmp_path):
    """Der eigentliche Zweck: in der GA-Funktionsliste darf geloescht werden."""
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    name = [n for n in buch.sheetnames if n.startswith("Übersicht")][0]
    liste = buch[[n for n in buch.sheetnames if n.startswith("GA-Funktionsliste")][0]]

    gesamt_zelle = (f"{get_column_letter(buch[name].max_column - 1)}"
                    f"{buch[name].max_row - 2}")          # Summenzeile, Spalte "Funktionen gesamt"
    vorher = float(_rechne(ziel)[(name.upper(), gesamt_zelle)])

    # Erste Datenzeile entfernen - so wie es in Excel jemand tun wuerde.
    entfallen = sum(wert for spalte in range(_erste_datenspalte(liste), liste.max_column - 1)
                    if isinstance(wert := liste.cell(6, spalte).value, (int, float)))
    assert entfallen > 0, "die geloeschte Zeile muss Mengen enthalten"
    letzte_datenzeile = liste.max_row - 3
    liste.delete_rows(letzte_datenzeile + 1, 3)   # Summe/Einheitspreis/Kosten der Liste
    liste.delete_rows(6)
    gekuerzt = tmp_path / "gekuerzt.xlsx"
    buch.save(gekuerzt)

    nachher = float(_rechne(gekuerzt)[(name.upper(), gesamt_zelle)])
    assert nachher == pytest.approx(vorher - entfallen)


@pytest.fixture(scope="module")
def zwei_listen(tmp_path_factory, samples):
    """Zwei Listen derselben Fassung - und zwar mit demselben Dateinamen.

    Genau hier muss sich zeigen, ob die Uebersicht die richtigen Zeilen der
    GA-Funktionsliste erwischt: ueber den Dateinamen waeren beide nicht zu
    unterscheiden, und ein Bezug auf feste Zeilenbereiche wuerde bei der
    zweiten Liste danebengreifen.
    """
    verzeichnis = tmp_path_factory.mktemp("zwei")
    engine = db.make_engine(verzeichnis / "test.sqlite3")
    with Session(engine) as session:
        erste = process_file(samples["alt"])
        db.save_document(session, erste)
        # Gleicher Dateiname, anderer Inhalt: nur jede zweite Zeile.
        zweite = copy.deepcopy(erste)
        zweite.file_hash = "zweite" + erste.file_hash[6:]
        zweite.rows = [zeile for nr, zeile in enumerate(zweite.rows) if nr % 2 == 0]
        db.save_document(session, zweite)

        roh = aggregate.raw_dataframe(session)
        matrizen = {p: aggregate.datei_spalten_matrix(session, p)
                    for p in aggregate.profiles_in_use(session)}
        ziel = export_excel.export_workbook(
            verzeichnis / "Auswertung.xlsx",
            summary=aggregate.column_summary(roh),
            raw=roh,
            documents=aggregate.documents_frame(session),
            footnotes=aggregate.footnotes_frame(session),
            prices=PREISE,
            layouts={p: aggregate.vdi_layout_frame(session, p)
                     for p in aggregate.profiles_in_use(session)},
            matrizen=matrizen,
        )
    return ziel, matrizen


def test_jede_zeile_zieht_nur_die_zeilen_ihrer_eigenen_liste(zwei_listen):
    """Zelle fuer Zelle gegen die Datenbank - je Datei und je Funktionsspalte."""
    ziel, matrizen = zwei_listen
    werte = _rechne(ziel)
    buch = openpyxl.load_workbook(ziel)
    name = [n for n in buch.sheetnames if n.startswith("Übersicht")][0]
    blatt = buch[name]

    erste = _erste_datenspalte(blatt)

    geprueft = 0
    for profil, matrix in matrizen.items():
        profile = get_profile(profil)
        datensaetze = matrix.to_dict("records")
        assert len(datensaetze) == 2, "der Fall lebt von zwei Listen"
        assert datensaetze[0]["Datei"] == datensaetze[1]["Datei"], \
            "gleicher Dateiname - unterschieden wird ueber den Schluessel"
        assert datensaetze[0]["Schlüssel"] != datensaetze[1]["Schlüssel"]
        unterschiedlich = [c.key for c in profile.columns
                           if datensaetze[0].get(c.key) != datensaetze[1].get(c.key)]
        assert unterschiedlich, \
            "die Listen muessen sich unterscheiden, sonst prueft der Test nichts"

        for versatz, record in enumerate(datensaetze):
            zeile = 6 + versatz
            assert blatt.cell(zeile, 1).value == record["Datei"]
            datenpunkte = werte[(name.upper(),
                                 f"{get_column_letter(erste - 1)}{zeile}")]
            assert float(datenpunkte) == float(record["Datenpunkte"]), \
                f"Datenpunkte in Zeile {zeile}"
            for index, column in enumerate(profile.columns):
                zelle = f"{get_column_letter(erste + index)}{zeile}"
                erwartet = float(record.get(column.key, 0.0) or 0.0)
                assert float(werte[(name.upper(), zelle)]) == pytest.approx(erwartet), \
                    f"Spalte {column.address} in Zeile {zeile}"
                geprueft += 1
    assert geprueft >= 50, "es muessen alle Funktionsspalten beider Zeilen geprueft sein"


def _rechne(pfad) -> dict:
    """Wertet alle Formeln der Mappe aus - Schluessel: (BLATT, "A1")."""
    formulas = pytest.importorskip("formulas", reason="Formelrechner nicht installiert")
    import logging
    import warnings

    warnings.filterwarnings("ignore")
    logging.disable(logging.WARNING)

    loesung = formulas.ExcelModel().loads(str(pfad)).finish().calculate()
    werte = {}
    for schluessel, wert in loesung.items():
        teile = schluessel.split("!")
        if len(teile) < 2:
            continue
        try:
            werte[(teile[-2].split("]")[-1].strip("'").upper(), teile[-1].strip("'"))] = wert.value[0, 0]
        except Exception:
            continue
    return werte


def _zahl(werte: dict, blatt: str, spalte: int, zeile: int) -> float:
    try:
        return float(werte[(blatt.upper(), f"{get_column_letter(spalte)}{zeile}")])
    except (KeyError, TypeError, ValueError):
        return 0.0


# --------------------------------------------------------------------------
# Alle Blaetter rechnen - und nur die wichtigen stehen vorn
# --------------------------------------------------------------------------

def test_nebenblaetter_sind_ausgeblendet(export):
    """Zwoelf Reiter sind zu viel - die Nebenblaetter bleiben im Hintergrund."""
    from vdi3814 import export_excel as modul

    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)

    for name in modul.NEBENBLAETTER:
        assert name in buch.sheetnames, f"'{name}' darf nicht fehlen, nur ausgeblendet sein"
        assert buch[name].sheet_state == "hidden", name

    sichtbar = [n for n in buch.sheetnames if buch[n].sheet_state == "visible"]
    assert any(n.startswith("Übersicht") for n in sichtbar)
    assert {"Schwerpunkte", "Kostenschätzung", "Prüfung"} <= set(sichtbar)
    assert any(n.startswith("GA-Funktionsliste") for n in sichtbar)
    # Eine Mappe, die auf einem ausgeblendeten Blatt aufschlaegt, oeffnet nicht
    assert buch.active.sheet_state == "visible"


def test_jedes_blatt_rechnet_mit_formeln(export):
    """Kein Blatt darf reine Kopien enthalten - sonst rechnet es nicht mit."""
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    ohne = [name for name in buch.sheetnames
            if not any(isinstance(zelle.value, str) and zelle.value.startswith("=")
                       for zeile in buch[name].iter_rows() for zelle in zeile)]
    assert not ohne, f"Blaetter ohne jede Formel: {ohne}"


def test_schwerpunkte_werden_gezaehlt_statt_kopiert(export):
    """Das Blatt "Schwerpunkte" muss an der Funktionsliste haengen."""
    ziel, mengen = export
    buch = openpyxl.load_workbook(ziel)
    blatt = buch["Schwerpunkte"]
    kopf = [zelle.value for zelle in blatt[1]]
    datenpunkte = get_column_letter(kopf.index("Datenpunkte") + 1)
    funktionen = get_column_letter(kopf.index("Funktionen") + 1)

    assert str(blatt[f"{datenpunkte}2"].value).startswith("=SUMIFS(")
    assert str(blatt[f"{funktionen}2"].value).startswith("=SUMIFS(")
    # Verknuepft wird ueber eine ausgeblendete Spalte, nicht ueber den Namen
    assert blatt.cell(1, blatt.max_column).value == "Schwerpunkt-Schlüssel"
    assert blatt.column_dimensions[get_column_letter(blatt.max_column)].hidden

    werte = _rechne(ziel)
    erwartet = _schwerpunkte_aus_rohdaten(buch)
    geprueft = 0
    for zeile in range(2, blatt.max_row):            # letzte Zeile ist die Summe
        kennung = blatt.cell(zeile, 1).value
        gerechnet = werte[("SCHWERPUNKTE", f"{funktionen}{zeile}")]
        assert float(gerechnet) == pytest.approx(erwartet[kennung]), kennung
        geprueft += 1
    assert geprueft, "es muss mindestens ein Schwerpunkt ausgewiesen sein"

    summe = werte[("SCHWERPUNKTE", f"{funktionen}{blatt.max_row}")]
    assert float(summe) == pytest.approx(sum(mengen.values()))


def test_mengen_je_schwerpunkt_zaehlen_die_zeilen_ihres_asp(export):
    """Auch die VDI-Sicht je Schwerpunkt zaehlt, statt Werte zu uebernehmen."""
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    name = [n for n in buch.sheetnames if n.startswith("Mengen je Schwerpunkt")][0]
    blatt = buch[name]
    erste = _erste_datenspalte(blatt)

    menge = str(blatt.cell(6, erste).value)
    assert menge.startswith("=SUMIFS("), menge
    assert "GA-Funktionsliste" in menge
    assert str(blatt.cell(6, erste - 1).value).startswith("=COUNTIFS(")

    # Zeilensumme je Schwerpunkt gegen die Einzelspalten
    werte = _rechne(ziel)
    gesamt_spalte = get_column_letter(blatt.max_column - 2)   # davor: zwei Schluessel
    zeile = 6
    while blatt.cell(zeile, 1).value and blatt.cell(zeile, 1).value != "Summe Funktionen":
        einzeln = sum(_zahl(werte, name, spalte, zeile)
                      for spalte in range(erste, blatt.max_column - 2))
        gesamt = werte[(name.upper(), f"{gesamt_spalte}{zeile}")]
        assert float(gesamt) == pytest.approx(einzeln), f"Zeile {zeile}"
        zeile += 1
    assert zeile > 6


def test_dokumente_und_info_verweisen_auf_die_uebersicht(export):
    """Auch die Nebenblaetter zeigen den Stand nach einer Aenderung."""
    ziel, mengen = export
    buch = openpyxl.load_workbook(ziel)
    dokumente = buch["Dokumente"]
    kopf = [zelle.value for zelle in dokumente[1]]
    summe_spalte = get_column_letter(kopf.index("summe_funktionen") + 1)
    assert str(dokumente[f"{summe_spalte}2"].value).startswith("='Übersicht")

    werte = _rechne(ziel)
    je_datei = [float(werte[("DOKUMENTE", f"{summe_spalte}{zeile}")])
                for zeile in range(2, dokumente.max_row)]
    assert sum(je_datei) == pytest.approx(sum(mengen.values()))

    info = buch["Info"]
    zeilen = {info.cell(z, 1).value: info.cell(z, 2).value for z in range(1, info.max_row + 1)}
    assert str(zeilen["Funktionen gesamt (Stand Export)"]).startswith("=")
    assert float(werte[("INFO", "B5")]) == pytest.approx(sum(mengen.values()))


def test_geloeschte_zeile_wirkt_bis_in_die_schwerpunkte(export, tmp_path):
    """Eine geloeschte Zeile muss auch in der ASP-Auswertung fehlen."""
    ziel, _ = export
    buch = openpyxl.load_workbook(ziel)
    liste = buch[[n for n in buch.sheetnames if n.startswith("GA-Funktionsliste")][0]]
    blatt = buch["Schwerpunkte"]
    kopf = [zelle.value for zelle in blatt[1]]
    funktionen = f"{get_column_letter(kopf.index('Funktionen') + 1)}{blatt.max_row}"

    vorher = float(_rechne(ziel)[("SCHWERPUNKTE", funktionen)])
    # Zwei Schluesselspalten stehen rechts, davor die Bemerkungen.
    entfallen = sum(wert for spalte in range(_erste_datenspalte(liste), liste.max_column - 2)
                    if isinstance(wert := liste.cell(6, spalte).value, (int, float)))
    assert entfallen > 0, "die geloeschte Zeile muss Mengen enthalten"

    letzte_datenzeile = liste.max_row - 3
    liste.delete_rows(letzte_datenzeile + 1, 3)      # Summe/Einheitspreis/Kosten
    liste.delete_rows(6)
    gekuerzt = tmp_path / "gekuerzt.xlsx"
    buch.save(gekuerzt)

    nachher = float(_rechne(gekuerzt)[("SCHWERPUNKTE", funktionen)])
    assert nachher == pytest.approx(vorher - entfallen)


def _schwerpunkte_aus_rohdaten(buch) -> dict:
    """Funktionen je Schwerpunkt - unabhaengig aus dem Blatt "Rohdaten"."""
    roh = buch["Rohdaten"]
    kopf = [zelle.value for zelle in roh[1]]
    kennung_spalte = kopf.index("schwerpunkt") + 1
    wert_spalte = kopf.index("wert") + 1
    summen: dict[str, float] = {}
    for zeile in range(2, roh.max_row):              # letzte Zeile ist die Summe
        kennung = roh.cell(zeile, kennung_spalte).value
        wert = roh.cell(zeile, wert_spalte).value
        summen[kennung] = summen.get(kennung, 0.0) + float(wert or 0.0)
    return summen
