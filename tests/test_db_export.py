"""Datenbank, Aggregation, Kosten und Excel-Export."""

import openpyxl
import pytest
from sqlalchemy.orm import Session

from vdi3814 import aggregate, db, export_excel
from vdi3814.costs import apply_prices, price_template, total_cost
from vdi3814.pipeline import process_file


@pytest.fixture()
def session(tmp_path):
    engine = db.make_engine(tmp_path / "test.sqlite3")
    with Session(engine) as active:
        yield active


def test_speichern_und_zurueckladen(session, samples):
    result = process_file(samples["alt"])
    document = db.save_document(session, result)
    reloaded = db.load_document(session, document.id)

    assert reloaded.fassung == "alt"
    assert len(reloaded.data_rows()) == len(result.data_rows())
    assert reloaded.grand_total() == result.grand_total()
    assert reloaded.footnotes and reloaded.pages


def test_erneuter_import_ersetzt_statt_zu_verdoppeln(session, samples):
    result = process_file(samples["alt"])
    db.save_document(session, result)
    document = db.save_document(session, result)

    assert len(db.list_documents(session)) == 1
    reloaded = db.load_document(session, document.id)
    assert reloaded.grand_total() == result.grand_total()


def test_inkrementell_erweiterbar(session, samples):
    db.save_document(session, process_file(samples["alt"]))
    db.save_document(session, process_file(samples["neu"]))
    assert len(db.list_documents(session)) == 2

    frame = aggregate.raw_dataframe(session)
    assert set(frame["fassung"]) == {"alt", "neu"}
    assert frame["wert"].sum() == 61          # 31 (alt) + 30 (neu)


def test_summenzeile_wird_nicht_mitgezaehlt(session, samples):
    db.save_document(session, process_file(samples["alt"]))
    frame = aggregate.raw_dataframe(session)
    assert frame["wert"].sum() == 31
    assert "Summe Funktionen" not in set(frame["datenpunkt"])


def test_dokumentuebersicht_protokolliert_uebersprungene_seiten(session, samples):
    db.save_document(session, process_file(samples["alt"]))
    documents = aggregate.documents_frame(session)
    assert documents.loc[0, "seiten"] == 2
    assert documents.loc[0, "seiten_ausgewertet"] == 1
    assert "schema" in documents.loc[0, "seiten_uebersprungen"]


def test_aufschluesselung_und_kosten(session, samples):
    db.save_document(session, process_file(samples["alt"]))
    frame = aggregate.raw_dataframe(session)
    summary = aggregate.column_summary(frame)

    gruppen = aggregate.group_summary(frame)
    assert gruppen["menge"].sum() == 31
    assert not aggregate.by_dimension(frame, "anlage").empty

    prices = {"A_1_5": 100.0, "A_1_1": 50.0}
    priced = apply_prices(summary, prices)
    erwartet = sum(
        row.menge * prices.get(row.spalte_key, 0.0) for row in summary.itertuples()
    )
    assert total_cost(priced) == pytest.approx(erwartet)


def test_preisvorlage_auch_ohne_import():
    template = price_template("vdi3814_neu")
    assert len(template) == 56
    assert set(template.columns) >= {"spalte_key", "spalte_adresse", "einheitspreis"}


def test_excel_export_enthaelt_formeln(session, samples, tmp_path):
    db.save_document(session, process_file(samples["alt"]))
    db.save_document(session, process_file(samples["neu"]))
    frame = aggregate.raw_dataframe(session)
    target = tmp_path / "auswertung.xlsx"
    export_excel.export_workbook(
        target,
        summary=aggregate.column_summary(frame),
        raw=frame,
        documents=aggregate.documents_frame(session),
        footnotes=aggregate.footnotes_frame(session),
        prices={"A_1_5": 80.0},
        projects=aggregate.pivot_projects(frame),
        matrizen={profil: aggregate.datei_spalten_matrix(session, profil)
                  for profil in aggregate.profiles_in_use(session)},
    )

    book = openpyxl.load_workbook(target)
    assert {"Kostenschätzung", "Spaltensummen", "Rohdaten", "Dokumente",
            "Fußnoten"} <= set(book.sheetnames)
    assert any(name.startswith("Übersicht") for name in book.sheetnames)

    costs = book["Kostenschätzung"]
    # Die Menge verweist auf die Summenzeile der Uebersicht - eine Quelle,
    # keine Kopie.
    assert str(costs.cell(2, 6).value).startswith("='Übersicht")
    assert costs.cell(2, 8).value == "=F2*G2"                        # Kosten als Formel
    assert str(costs.cell(costs.max_row, 8).value).startswith("=SUM(")

    spaltensummen = book["Spaltensummen"]
    assert spaltensummen.cell(spaltensummen.max_row, 1).value == "Summe Funktionen"

    raw_sheet = book["Rohdaten"]
    assert "datei" in [c.value for c in raw_sheet[1]]                # Quelldatei-Referenz


# --------------------------------------------------------------------------
# Gespeicherter Datensatz bleibt nach dem Schliessen der Sitzung lesbar
# --------------------------------------------------------------------------

def test_gespeichertes_dokument_bleibt_nach_dem_schliessen_lesbar(tmp_path, samples):
    """Die Oberflaeche liest die Nummer, wenn die Sitzung schon zu ist.

    Genau daran ist das Speichern in Reiter 2 gescheitert: nach dem commit
    waren alle Felder als veraltet markiert, und beim Zugriff nach dem
    Schliessen kam ein DetachedInstanceError statt einer Erfolgsmeldung.
    """
    engine = db.make_engine(tmp_path / "test.sqlite3")
    with Session(engine) as session:
        dokument = db.save_document(session, process_file(samples["alt"]))

    assert isinstance(dokument.id, int)
    assert dokument.file_name.endswith(".pdf")


# --------------------------------------------------------------------------
# Abgelegte Kostenschaetzungen
# --------------------------------------------------------------------------

POSITIONEN = [
    {"spalte_adresse": "1.1", "gruppe": "Ein-/Ausgabefunktionen", "spalte": "BI",
     "spalte_key": "A_1_1", "menge": 10.0, "einheitspreis": 45.0},
    {"spalte_adresse": "1.5", "gruppe": "Ein-/Ausgabefunktionen", "spalte": "AO",
     "spalte_key": "A_1_5", "menge": 4.0, "einheitspreis": 80.0},
]


def test_kostenschaetzung_ablegen_und_wieder_aufrufen(session):
    nummer = db.save_cost_estimate(session, "Angebot", POSITIONEN, bemerkung="Erstkalkulation")

    abgelegt = db.list_cost_estimates(session)
    assert [e["name"] for e in abgelegt] == ["Angebot"]
    assert abgelegt[0]["positionen"] == 2
    assert abgelegt[0]["kosten_gesamt"] == pytest.approx(10 * 45.0 + 4 * 80.0)
    assert abgelegt[0]["bemerkung"] == "Erstkalkulation"

    stand = db.load_cost_estimate(session, nummer)
    assert [p["spalte_key"] for p in stand["positionen"]] == ["A_1_1", "A_1_5"]
    assert stand["positionen"][0]["kosten"] == pytest.approx(450.0)
    assert stand["menge_gesamt"] == pytest.approx(14.0)


def test_gleicher_name_schreibt_den_stand_fort(session):
    db.save_cost_estimate(session, "Angebot", POSITIONEN)
    teurer = [{**POSITIONEN[0], "einheitspreis": 60.0}]
    db.save_cost_estimate(session, "Angebot", teurer)

    abgelegt = db.list_cost_estimates(session)
    assert len(abgelegt) == 1, "derselbe Name darf keine Dublette anlegen"
    assert abgelegt[0]["kosten_gesamt"] == pytest.approx(600.0)


def test_kostenschaetzung_laesst_sich_loeschen(session):
    nummer = db.save_cost_estimate(session, "Variante", POSITIONEN)
    db.delete_cost_estimate(session, nummer)

    assert db.list_cost_estimates(session) == []
    assert db.load_cost_estimate(session, nummer) is None
    # Die Positionen duerfen nicht als Waisen zurueckbleiben
    assert session.query(db.CostEstimateItem).count() == 0


def test_ohne_namen_bekommt_der_stand_einen_zeitstempel(session):
    nummer = db.save_cost_estimate(session, "   ", POSITIONEN)
    stand = db.load_cost_estimate(session, nummer)
    assert stand["name"].startswith("Stand ")
